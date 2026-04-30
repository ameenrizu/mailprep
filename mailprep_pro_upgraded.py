

import os
import re
import sys
import csv
import json
import html
import zipfile
from dataclasses import dataclass
from datetime import datetime
from typing import List, Tuple, Optional, Dict

from PyQt5.QtCore import Qt, QEvent, QThread, pyqtSignal, QObject, QTimer, QMimeData
from PyQt5.QtGui import QClipboard, QGuiApplication, QKeySequence, QFont
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QFileDialog,
    QMessageBox,
    QLabel,
    QPushButton,
    QTextEdit,
    QLineEdit,
    QHBoxLayout,
    QVBoxLayout,
    QFrame,
    QProgressBar,
    QCheckBox,
    QGridLayout,
    QToolButton,
    QShortcut,
    QDialog,
    QScrollArea,
    QAction,
    QMenu,
    QComboBox,
)

# ------------------------------------------------------------
# Debug flag — set MAILPREP_DEBUG=1 in environment to enable
# ------------------------------------------------------------
DEBUG = os.environ.get("MAILPREP_DEBUG") == "1"

# ------------------------------------------------------------
# Constants
# ------------------------------------------------------------
MARKER_FOLDERS = {
    "exr",
    "support_files",
    "paint",
    "comp",
    "roto",
    "plates",
    "renders",
    "delivery",
    "nuke",
    "_exr",
    "_support_files",
    "_paint",
    "_comp",
    "_roto",
    "_plates",
    "_renders",
    "_delivery",
    "_nuke",
}

SEQUENCE_EXTENSIONS_HINT = {
    ".exr",
    ".dpx",
    ".png",
    ".jpg",
    ".jpeg",
    ".tif",
    ".tiff",
}

KNOWN_FILE_EXTENSIONS = {
    ".nk",
    ".ma",
    ".mb",
    ".fbx",
    ".abc",
    ".obj",
    ".usd",
    ".usda",
    ".usdc",
    ".mov",
    ".mp4",
    ".mxf",
    ".avi",
    ".wav",
    ".mp3",
    ".aif",
    ".aiff",
    ".flac",
    ".jpg",
    ".jpeg",
    ".png",
    ".tif",
    ".tiff",
    ".exr",
    ".dpx",
    ".gif",
    ".bmp",
    ".psd",
    ".txt",
    ".csv",
    ".json",
    ".xml",
    ".yml",
    ".yaml",
    ".pdf",
    ".zip",
    ".rar",
    ".7z",
    ".sfx",
    ".ccc",
    ".cdl",
    ".cube",
    ".lut",
}

# ------------------------------------------------------------
# Ignored submission parent names — edit here to add new ones
# ------------------------------------------------------------
IGNORED_SUBMISSION_EXACT = {
    "h264",
    "h.264",
    "h265",
    "hevc",
    "mp4",
    "dnxhd",
    "dnxhr",
    "dnx",
    "prores",
    "pro_res",
    "proreshq",
    "avid",
    "mov",
    "mxf",
    "exr",
    "_exr",
    "jpeg",
    "jpg",
    "png",
    "tif",
    "tiff",
    "tif16",
    "tiff16",
    "delivery",
    "deliveries",
    "support_files",
    "_support_files",
    "support",
    "lut",
    "luts",
    "cdl",
    "ccc",
    "plates",
    "plate",
    "renders",
    "render",
    "output",
    "outputs",
    "preview",
    "previews",
    "qt",
    "quicktime",
    "review",
    "client",
    "publish",
    "published",
    "final",
    "temp",
    "export",
    "exports",
    "web",
    "maya",
    "nuke",
    "script",
    "scripts",
    "splines",
    "rotomation",
    "dailies",
    "fbx",
    "undistorted_plate",
    "holdout",
    "perspective",
    "perspective2",
    "shaded",
    "wireframe",
    "curves",
    "perspectivestab",
    "pointblast",
    "pointblastdigorychest",
}

IGNORED_SUBMISSION_FUZZY = (
    "h264",
    "h265",
    "hevc",
    "dnx",
    "prores",
    "quicktime",
    "qt",
    "review",
    "delivery",
    "deliver",
    "output",
    "export",
    "publish",
    "mov",
    "mp4",
    "mxf",
    "avid",
    "web",
    "support",
    "lut",
    "ccc",
    "cdl",
)

REMOVABLE_VERSION_SUFFIXES = [
    "h264",
    "h265",
    "hevc",
    "dnxhd",
    "dnxhr",
    "dnx",
    "prores",
    "pro_res",
    "proreshq",
    "qt",
    "quicktime",
    "mov",
    "mxf",
    "mp4",
    "avid",
    "review",
    "web",
    "final",
    "output",
    "export",
    "vfx",
    "rotoslap",
    "slap",
    "script",
    "sfx",
    "nk",
    "mb",
    "ma",
    "fbx",
    "jpg",
    "jpeg",
    "png",
    "tif",
    "tiff",
    "exr",
    "ccc",
    "cdl",
]

SAFE_NAME_PATTERN = re.compile(r"^[A-Za-z0-9._-]+$")
RECENT_PATHS_FILE = os.path.expanduser("~/.mailprep_recent.json")
MAX_RECENT_PATHS = 8
TEMPLATE_SHEET_EXTENSIONS = {".csv", ".xlsx", ".xls"}

CURRENT_THEME = "dark"

THEMES = {
    "dark": {
        "root_bg": "#0d0f14",
        "panel_bg": "#151821",
        "toolbar_bg": "#11141a",
        "status_bg": "#11141a",
        "text_bg": "#0a0c10",
        "text_fg": "#f5f7fa",
        "file_fg": "#d7dce2",
        "folder_fg": "#ffffff",
        "missing_fg": "#ffb86b",
        "muted_fg": "#aeb6c2",
        "entry_bg": "#0b0d12",
        "entry_fg": "#f5f7fa",
        "border": "#2d3340",
        "button_bg": "#4f7cff",
        "button_fg": "#ffffff",
        "button_alt_bg": "#232833",
        "button_alt_fg": "#f5f7fa",
        "copy_rich_bg": "#2f8f5b",
        "copy_rich_fg": "#ffffff",
        "copy_rich_border": "#49b879",
        "copy_html_bg": "#b23b3b",
        "copy_html_fg": "#ffffff",
        "copy_html_border": "#df6c6c",
        "copy_full_bg": "#6b46c1",
        "copy_full_fg": "#ffffff",
        "copy_full_border": "#8b6be6",
        "copy_notes_bg": "#0f8aa8",
        "copy_notes_fg": "#ffffff",
        "copy_notes_border": "#46b4d4",
        "preview_count_bg": "#2f5fa7",
        "preview_count_fg": "#ffffff",
        "progress_bg": "#4f7cff",
        "section_title_fg": "#dfe7f7",
        "preview_header_bg": "#121722",
        "preview_header_border": "#2b3342",
        "primary_button_bg": "#4f7cff",
        "primary_button_hover": "#6790ff",
        "suspicious_bg": "#5e3b00",
        "suspicious_fg": "#ffd88a",
        "suspicious_border": "#d39c2f",
        "duplicate_bg": "#4a1f3d",
        "duplicate_fg": "#ffb3da",
        "duplicate_border": "#c85f95",
        "valid_bg": "#163a27",
        "valid_fg": "#9be8b0",
        "valid_border": "#2f8f5b",
        "warn_bg": "#3f2b11",
        "warn_fg": "#ffd68f",
        "warn_border": "#b9832f",
        "error_bg": "#451c24",
        "error_fg": "#ffb8c4",
        "error_border": "#b85b72",
    },
    "light": {
        "root_bg": "#eef2f7",
        "panel_bg": "#ffffff",
        "toolbar_bg": "#f4f7fb",
        "status_bg": "#f4f7fb",
        "text_bg": "#ffffff",
        "text_fg": "#111827",
        "file_fg": "#374151",
        "folder_fg": "#111827",
        "missing_fg": "#b45309",
        "muted_fg": "#667085",
        "entry_bg": "#ffffff",
        "entry_fg": "#111827",
        "border": "#d2d9e3",
        "button_bg": "#2563eb",
        "button_fg": "#ffffff",
        "button_alt_bg": "#eef2f7",
        "button_alt_fg": "#111827",
        "copy_rich_bg": "#2f8f5b",
        "copy_rich_fg": "#ffffff",
        "copy_rich_border": "#4caf7c",
        "copy_html_bg": "#c94848",
        "copy_html_fg": "#ffffff",
        "copy_html_border": "#e17878",
        "copy_full_bg": "#7c3aed",
        "copy_full_fg": "#ffffff",
        "copy_full_border": "#a78bfa",
        "copy_notes_bg": "#0284c7",
        "copy_notes_fg": "#ffffff",
        "copy_notes_border": "#7dd3fc",
        "preview_count_bg": "#dbeafe",
        "preview_count_fg": "#1d4ed8",
        "progress_bg": "#2563eb",
        "section_title_fg": "#334155",
        "preview_header_bg": "#f8fafc",
        "preview_header_border": "#d8dee8",
        "primary_button_bg": "#2563eb",
        "primary_button_hover": "#3b82f6",
        "suspicious_bg": "#fff3cd",
        "suspicious_fg": "#8a5300",
        "suspicious_border": "#e3b341",
        "duplicate_bg": "#fde7f3",
        "duplicate_fg": "#9d174d",
        "duplicate_border": "#f472b6",
        "valid_bg": "#eafaf0",
        "valid_fg": "#166534",
        "valid_border": "#4caf7c",
        "warn_bg": "#fff7e7",
        "warn_fg": "#9a6700",
        "warn_border": "#e3b341",
        "error_bg": "#fff1f2",
        "error_fg": "#be123c",
        "error_border": "#fb7185",
    },
}


# ------------------------------------------------------------
# Data models
# ------------------------------------------------------------
@dataclass
class LineItem:
    item_type: str  # folder | file | missing | blank
    level: int
    text: str


@dataclass
class BuildResult:
    lines: List[LineItem]
    plain_text: str
    html_text: str
    clipboard_html: str
    folder_count: int
    file_count: int
    signature: Optional[Tuple]


# ------------------------------------------------------------
# Core logic
# ------------------------------------------------------------
class MailPrepLogic:
    @staticmethod
    def is_suspicious_name(name: str) -> bool:
        name = (name or "").strip()
        if name.endswith(" (contains)"):
            name = name[: -len(" (contains)")].rstrip()
        if not name:
            return False
        return SAFE_NAME_PATTERN.fullmatch(name) is None

    @staticmethod
    def group_sequences(files: List[str]) -> List[str]:
        pattern = re.compile(r"^(.*?)(\d+)(\.[^.]+)$")
        grouped: Dict[Tuple[str, str], List[Tuple[int, int]]] = {}
        non_sequence = []

        for f in files:
            match = pattern.match(f)
            if match:
                base, num, ext = match.groups()
                key = (base, ext)
                grouped.setdefault(key, []).append((int(num), len(num)))
            else:
                non_sequence.append(f)

        result = []

        for (base, ext), items in grouped.items():
            num_to_width = {}
            for num, width in items:
                num_to_width[num] = max(width, num_to_width.get(num, 0))

            sorted_nums = sorted(num_to_width.keys())
            if not sorted_nums:
                continue

            max_width = max(num_to_width.values())
            start = sorted_nums[0]
            prev = sorted_nums[0]

            for current in sorted_nums[1:]:
                if current == prev + 1:
                    prev = current
                else:
                    if start == prev:
                        result.append(f"{base}{str(start).zfill(max_width)}{ext}")
                    else:
                        result.append(
                            f"{base}{str(start).zfill(max_width)}-{str(prev).zfill(max_width)}{ext}"
                        )
                    start = current
                    prev = current

            if start == prev:
                result.append(f"{base}{str(start).zfill(max_width)}{ext}")
            else:
                result.append(
                    f"{base}{str(start).zfill(max_width)}-{str(prev).zfill(max_width)}{ext}"
                )

        result.extend(sorted(non_sequence))
        return sorted(result)

    @staticmethod
    def _compress_number_ranges(numbers: List[int], width: int) -> str:
        if not numbers:
            return ""

        numbers = sorted(set(numbers))
        parts = []
        start = numbers[0]
        prev = numbers[0]

        for n in numbers[1:]:
            if n == prev + 1:
                prev = n
            else:
                if start == prev:
                    parts.append(str(start).zfill(width))
                else:
                    parts.append(f"{str(start).zfill(width)}-{str(prev).zfill(width)}")
                start = prev = n

        if start == prev:
            parts.append(str(start).zfill(width))
        else:
            parts.append(f"{str(start).zfill(width)}-{str(prev).zfill(width)}")

        return ", ".join(parts)

    @staticmethod
    def detect_missing_ranges(files: List[str]) -> List[str]:
        pattern = re.compile(r"^(.*?)(\d+)(\.[^.]+)$")
        grouped: Dict[Tuple[str, str], List[Tuple[int, int]]] = {}

        for f in files:
            match = pattern.match(f)
            if not match:
                continue

            base, num, ext = match.groups()
            if ext.lower() not in SEQUENCE_EXTENSIONS_HINT and len(num) < 3:
                continue

            key = (base, ext)
            grouped.setdefault(key, []).append((int(num), len(num)))

        missing_lines = []

        for (base, ext), items in grouped.items():
            unique = {}
            for num, width in items:
                unique[num] = max(width, unique.get(num, 0))

            nums = sorted(unique.keys())
            if len(nums) < 2:
                continue

            width = max(unique.values())
            missing = []

            for i in range(len(nums) - 1):
                current_num = nums[i]
                next_num = nums[i + 1]
                if next_num - current_num > 1:
                    missing.extend(range(current_num + 1, next_num))

            if not missing:
                continue

            compressed = MailPrepLogic._compress_number_ranges(missing, width)
            label = f"Missing frames: {compressed}"

            if len(grouped) > 1:
                label = f"Missing frames ({base}*{ext}): {compressed}"

            missing_lines.append(label)

        return sorted(missing_lines)

    @staticmethod
    def get_dir_entries(path: str) -> Tuple[List[str], List[str]]:
        try:
            entries = sorted(os.listdir(path))
        except Exception:
            return [], []

        dirs = []
        files = []

        for name in entries:
            full = os.path.join(path, name)
            if os.path.isdir(full):
                dirs.append(name)
            else:
                files.append(name)

        return dirs, files

    @staticmethod
    def is_single_shot_root(path: str) -> bool:
        try:
            entries = os.listdir(path)
        except Exception:
            return False

        subdirs = {
            name.lower() for name in entries if os.path.isdir(os.path.join(path, name))
        }
        return any(name in MARKER_FOLDERS for name in subdirs)

    @staticmethod
    def should_skip_root_for_single_marker(path: str) -> bool:
        dirs, files = MailPrepLogic.get_dir_entries(path)
        if files:
            return False
        if len(dirs) != 1:
            return False
        return dirs[0].lower() in MARKER_FOLDERS

    @staticmethod
    def walk_tree_files_first(
        current_path: str, current_level: int, lines: List[LineItem]
    ) -> None:
        dirs, files = MailPrepLogic.get_dir_entries(current_path)

        MailPrepLogic.add_files_with_zips(files, current_path, current_level, lines)

        missing_lines = MailPrepLogic.detect_missing_ranges(files)
        for missing_text in missing_lines:
            lines.append(LineItem("missing", current_level, missing_text))

        for d in dirs:
            lines.append(LineItem("folder", current_level, d))
            MailPrepLogic.walk_tree_files_first(
                os.path.join(current_path, d),
                current_level + 1,
                lines,
            )

    @staticmethod
    def build_single_shot_lines(path: str, level: int = 0) -> List[LineItem]:
        lines: List[LineItem] = []
        root_name = os.path.basename(path.rstrip(os.sep))
        lines.append(LineItem("folder", level, root_name))
        MailPrepLogic.walk_tree_files_first(path, level + 1, lines)
        return lines

    @staticmethod
    def build_single_shot_contents_only(path: str, level: int = 0) -> List[LineItem]:
        lines: List[LineItem] = []
        MailPrepLogic.walk_tree_files_first(path, level, lines)
        return lines

    @staticmethod
    def build_shot_block(path: str, level: int = 0) -> List[LineItem]:
        lines: List[LineItem] = []
        shot_name = os.path.basename(path.rstrip(os.sep))
        lines.append(LineItem("folder", level, shot_name))
        MailPrepLogic.walk_tree_files_first(path, level + 1, lines)
        return lines

    @staticmethod
    def add_files_with_zips(
        files: List[str], current_path: str, current_level: int, lines: List[LineItem]
    ) -> None:
        grouped_files = MailPrepLogic.group_sequences(files)
        for f in grouped_files:
            if f.lower().endswith(".zip"):
                lines.append(LineItem("file", current_level, f"{f} (contains)"))
            else:
                lines.append(LineItem("file", current_level, f))

    @staticmethod
    def build_zip_contents(zip_path: str, level: int) -> List[LineItem]:
        """
        Builds a tree of LineItems from a zip file's contents.
        Reserved for future use (e.g. optional QC expand view).
        Not used in email/copy output — zips always show as '(contains)'.
        """
        lines = []
        try:
            with zipfile.ZipFile(zip_path, "r") as zf:
                namelist = zf.namelist()
                tree = {}
                for name in namelist:
                    if not name.endswith("/"):
                        parts = name.split("/")
                        current = tree
                        for part in parts[:-1]:
                            if part not in current:
                                current[part] = {}
                            current = current[part]
                        if parts[-1]:
                            current[parts[-1]] = None

                def walk_tree(node, current_level):
                    for key, value in sorted(node.items()):
                        if value is None:
                            lines.append(LineItem("file", current_level, key))
                        else:
                            lines.append(LineItem("folder", current_level, key))
                            walk_tree(value, current_level + 1)

                walk_tree(tree, level)
        except Exception as e:
            lines.append(
                LineItem("file", level, f"Error reading zip contents: {str(e)}")
            )
        return lines

    @staticmethod
    def build_package_lines(path: str) -> List[LineItem]:
        if MailPrepLogic.should_skip_root_for_single_marker(path):
            return MailPrepLogic.build_single_shot_contents_only(path, level=0)

        if MailPrepLogic.is_single_shot_root(path):
            return MailPrepLogic.build_single_shot_lines(path)

        root_dirs, root_files = MailPrepLogic.get_dir_entries(path)

        if len(root_dirs) == 1 and not root_files:
            only_child = os.path.join(path, root_dirs[0])

            if MailPrepLogic.should_skip_root_for_single_marker(only_child):
                lines = [LineItem("folder", 0, root_dirs[0])]
                child_lines = MailPrepLogic.build_single_shot_contents_only(
                    only_child, level=1
                )
                lines.extend(child_lines)
                return lines

            if MailPrepLogic.is_single_shot_root(only_child):
                return MailPrepLogic.build_single_shot_lines(only_child)

            lines: List[LineItem] = []
            package_name = root_dirs[0]
            lines.append(LineItem("folder", 0, package_name))

            package_dirs, package_files = MailPrepLogic.get_dir_entries(only_child)

            MailPrepLogic.add_files_with_zips(package_files, only_child, 1, lines)

            missing_lines = MailPrepLogic.detect_missing_ranges(package_files)
            for missing_text in missing_lines:
                lines.append(LineItem("missing", 1, missing_text))

            for idx, d in enumerate(package_dirs):
                shot_path = os.path.join(only_child, d)
                shot_block = MailPrepLogic.build_shot_block(shot_path, level=1)
                lines.extend(shot_block)
                if idx < len(package_dirs) - 1:
                    lines.append(LineItem("blank", 0, ""))

            return lines

        lines: List[LineItem] = []

        MailPrepLogic.add_files_with_zips(root_files, path, 0, lines)

        missing_lines = MailPrepLogic.detect_missing_ranges(root_files)
        for missing_text in missing_lines:
            lines.append(LineItem("missing", 0, missing_text))

        for idx, d in enumerate(root_dirs):
            shot_path = os.path.join(path, d)
            shot_block = MailPrepLogic.build_shot_block(shot_path, level=0)
            lines.extend(shot_block)
            if idx < len(root_dirs) - 1:
                lines.append(LineItem("blank", 0, ""))

        return lines

    @staticmethod
    def make_plain_text_output(lines: List[LineItem]) -> str:
        output = []
        for item in lines:
            if item.item_type == "blank":
                output.append("")
            else:
                indent = "    " * item.level
                output.append(f"{indent}{item.text}")
        return "\n".join(str(x) for x in output)

    @staticmethod
    def _format_preview_text(
        text: str,
        theme_name: str,
        suspicious: bool = False,
        item_type: str = "file",
    ) -> str:
        safe_text = html.escape(text)
        if not suspicious:
            return safe_text
        theme = THEMES[theme_name]
        return (
            f'<span style="background:{theme["error_bg"]}; '
            f'color:{theme["error_fg"]}; '
            f'border:1px solid {theme["error_border"]}; '
            f'border-radius:4px; padding:1px 4px;">'
            f"{safe_text}</span>"
        )

    @staticmethod
    def make_html_output(lines: List[LineItem], theme_name: str = "dark") -> str:
        theme = THEMES[theme_name]

        html_parts = [
            "<html><body>",
            (
                f'<div style="font-family:Calibri, Arial, sans-serif; '
                f'font-size:12pt; line-height:1.2; color:{theme["text_fg"]};">'
            ),
        ]

        for item in lines:
            if item.item_type == "blank":
                html_parts.append('<div style="height:0.45em;"></div>')
                continue

            indent = "&nbsp;" * (4 * item.level)
            is_suspicious = MailPrepLogic.is_suspicious_name(item.text)

            if item.item_type == "folder":
                formatted = MailPrepLogic._format_preview_text(
                    item.text,
                    theme_name,
                    suspicious=is_suspicious,
                    item_type="folder",
                )
                html_parts.append(
                    f'<div style="white-space:pre; color:{theme["folder_fg"]};">{indent}<b>{formatted}</b></div>'
                )
            elif item.item_type == "missing":
                safe_text = html.escape(item.text)
                html_parts.append(
                    f'<div style="white-space:pre;">{indent}'
                    f'<span style="background:{theme["error_bg"]}; '
                    f'color:{theme["error_fg"]}; '
                    f'border:1px solid {theme["error_border"]}; '
                    f'border-radius:4px; padding:1px 4px;">'
                    f"{safe_text}</span></div>"
                )
            else:
                formatted = MailPrepLogic._format_preview_text(
                    item.text,
                    theme_name,
                    suspicious=is_suspicious,
                    item_type="file",
                )
                html_parts.append(
                    f'<div style="white-space:pre; color:{theme["file_fg"]};">{indent}{formatted}</div>'
                )

        html_parts.append("</div></body></html>")
        return "\n".join(str(x) for x in html_parts)

    @staticmethod
    def make_clipboard_rich_html(lines: List[LineItem]) -> str:
        parts = [
            "<html><body>",
            '<div style="font-family:Calibri, Arial, sans-serif; '
            'font-size:12pt; line-height:1.2; color:#000000; background-color:transparent;">',
        ]

        for item in lines:
            if item.item_type == "blank":
                parts.append("<br>")
                continue

            indent = "&nbsp;" * (4 * item.level)
            safe_text = html.escape(item.text)
            content = f"{indent}{safe_text}"

            if item.item_type == "folder":
                content = f"<b>{content}</b>"

            parts.append(f"{content}<br>")

        parts.append("</div></body></html>")
        return "".join(str(x) for x in parts)

    @staticmethod
    def count_preview_items(lines: List[LineItem]) -> Tuple[int, int]:
        folder_count = sum(1 for item in lines if item.item_type == "folder")
        file_count = sum(1 for item in lines if item.item_type == "file")
        return folder_count, file_count

    @staticmethod
    def path_signature(path: str) -> Optional[Tuple]:
        try:
            root_stat = os.stat(path)
            sig_parts = [("ROOT", int(root_stat.st_mtime_ns), root_stat.st_size)]

            for dirpath, dirnames, filenames in os.walk(path):
                dirnames.sort()
                filenames.sort()

                rel_dir = os.path.relpath(dirpath, path)
                if rel_dir == ".":
                    rel_dir = ""

                for dirname in dirnames:
                    full = os.path.join(dirpath, dirname)
                    rel = os.path.join(rel_dir, dirname) if rel_dir else dirname
                    try:
                        stat = os.stat(full)
                        sig_parts.append(
                            (f"{rel}{os.sep}", int(stat.st_mtime_ns), stat.st_size)
                        )
                    except Exception:
                        sig_parts.append((f"{rel}{os.sep}", 0, 0))

                for filename in filenames:
                    full = os.path.join(dirpath, filename)
                    rel = os.path.join(rel_dir, filename) if rel_dir else filename
                    try:
                        stat = os.stat(full)
                        sig_parts.append((rel, int(stat.st_mtime_ns), stat.st_size))
                    except Exception:
                        sig_parts.append((rel, 0, 0))

            return tuple(sig_parts)
        except Exception:
            return None

    @staticmethod
    def build_result_from_lines(
        lines: List[LineItem],
        theme_name: str = "dark",
        signature: Optional[Tuple] = None,
    ) -> BuildResult:
        plain_text = MailPrepLogic.make_plain_text_output(lines)
        html_text = MailPrepLogic.make_html_output(lines, theme_name=theme_name)
        clipboard_html = MailPrepLogic.make_clipboard_rich_html(lines)
        folder_count, file_count = MailPrepLogic.count_preview_items(lines)

        return BuildResult(
            lines=lines,
            plain_text=plain_text,
            html_text=html_text,
            clipboard_html=clipboard_html,
            folder_count=folder_count,
            file_count=file_count,
            signature=signature,
        )

    @staticmethod
    def build_result_from_path(path: str, theme_name: str = "dark") -> BuildResult:
        lines = MailPrepLogic.build_package_lines(path)
        signature = MailPrepLogic.path_signature(path)
        return MailPrepLogic.build_result_from_lines(lines, theme_name, signature)

    @staticmethod
    def manifest_to_lines(manifest_data: dict) -> List[LineItem]:
        lines = []
        items = manifest_data.get("items", [])
        for item in items:
            try:
                lines.append(
                    LineItem(
                        item_type=str(item.get("type", "file")),
                        level=int(item.get("level", 0)),
                        text=str(item.get("text", "")),
                    )
                )
            except Exception:
                continue
        return lines

    @staticmethod
    def build_result_from_manifest(
        manifest_data: dict, theme_name: str = "dark"
    ) -> BuildResult:
        lines = MailPrepLogic.manifest_to_lines(manifest_data)
        signature = ("MANIFEST", len(lines), manifest_data.get("schema_version", "1.0"))
        return MailPrepLogic.build_result_from_lines(lines, theme_name, signature)

    @staticmethod
    def lines_to_manifest(lines: List[LineItem]) -> dict:
        return {
            "schema_version": "1.0",
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "items": [
                {"type": item.item_type, "level": item.level, "text": item.text}
                for item in lines
            ],
        }

    @staticmethod
    def build_result(path: str, theme_name: str = "dark") -> BuildResult:
        return MailPrepLogic.build_result_from_path(path, theme_name)


# ------------------------------------------------------------
# Worker
# ------------------------------------------------------------
class PreviewWorker(QObject):
    finished = pyqtSignal(object)
    failed = pyqtSignal(str)

    def __init__(self, input_mode: str, source_data, theme_name: str):
        super().__init__()
        self.input_mode = input_mode
        self.source_data = source_data
        self.theme_name = theme_name
        self._cancel = False

    def cancel(self):
        self._cancel = True

    def run(self):
        try:
            if self._cancel:
                return
            if self.input_mode == "path":
                result = MailPrepLogic.build_result_from_path(
                    self.source_data, self.theme_name
                )
            elif self.input_mode == "manifest":
                result = MailPrepLogic.build_result_from_manifest(
                    self.source_data, self.theme_name
                )
            else:
                raise ValueError(f"Unsupported input mode: {self.input_mode}")

            if not self._cancel:
                self.finished.emit(result)
        except Exception as exc:
            if not self._cancel:
                self.failed.emit(str(exc))


# ------------------------------------------------------------
# Preview Dialog
# ------------------------------------------------------------
class PreviewDialog(QDialog):
    def __init__(self, parent=None, html_content: str = "", theme_name: str = "dark"):
        super().__init__(parent)
        self.setWindowTitle("Preview")
        self.resize(900, 700)
        self.setMinimumSize(700, 500)

        theme = THEMES[theme_name]

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        header = QLabel("Mail Preview")
        header.setObjectName("dialogHeader")

        self.preview_box = QTextEdit()
        self.preview_box.setReadOnly(True)
        self.preview_box.setAcceptRichText(True)
        self.preview_box.setLineWrapMode(QTextEdit.NoWrap)
        self.preview_box.setHtml(html_content)

        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.accept)

        footer_row = QHBoxLayout()
        footer_row.addStretch(1)
        footer_row.addWidget(close_btn)

        layout.addWidget(header)
        layout.addWidget(self.preview_box, 1)
        layout.addLayout(footer_row)

        self.setStyleSheet(
            f"""
            QDialog {{
                background: {theme["root_bg"]};
            }}
            QLabel#dialogHeader {{
                color: {theme["text_fg"]};
                font-size: 18px;
                font-weight: 700;
                padding: 4px 2px;
            }}
            QTextEdit {{
                background: {theme["text_bg"]};
                color: {theme["text_fg"]};
                border: 1px solid {theme["border"]};
                border-radius: 10px;
                padding: 10px;
            }}
            QPushButton {{
                background: {theme["button_alt_bg"]};
                color: {theme["button_alt_fg"]};
                border: 1px solid {theme["border"]};
                border-radius: 10px;
                padding: 8px 14px;
                font-weight: 600;
            }}
            """
        )


class ZoomableTextEdit(QTextEdit):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._base_font_size = 12.0
        self._current_zoom_factor = 1.0
        self._raw_html = ""
        self.setFocusPolicy(Qt.ClickFocus)

    def _normalize_html(self, html_content: str) -> str:
        html_content = re.sub(r"(?is)^\s*<html.*?>", "", html_content)
        html_content = re.sub(r"(?is)</html>\s*$", "", html_content)
        html_content = re.sub(r"(?is)^\s*<body.*?>", "", html_content)
        html_content = re.sub(r"(?is)</body>\s*$", "", html_content)
        return html_content

    def _update_content(self):
        if not self._raw_html:
            return

        inner_html = self._normalize_html(self._raw_html)
        inner_html = re.sub(r"(?i)font-size\s*:\s*[^;\"\\)]+;?", "", inner_html)

        zoom_size = max(6.0, self._base_font_size * self._current_zoom_factor)
        css = (
            "html, body, div, span, p, li, td, th {"
            f" font-size: {zoom_size}pt !important;"
            " line-height: 1.2 !important; }"
        )
        self.document().setDefaultStyleSheet(css)
        super().setHtml(inner_html)
        self.viewport().update()

    def _apply_zoom_step(self, step: int):
        zoom_step = 0.1
        if step > 0:
            self._current_zoom_factor *= 1 + zoom_step * step
        elif step < 0:
            self._current_zoom_factor *= 1 - zoom_step * abs(step)
        self._update_content()

    def zoom_in_step(self):
        self._apply_zoom_step(1)

    def zoom_out_step(self):
        self._apply_zoom_step(-1)

    def reset_zoom(self):
        self._current_zoom_factor = 1.0
        self._update_content()

    def setHtml(self, html: str):
        self._raw_html = html or ""
        self._update_content()

    def wheelEvent(self, event):
        if event.modifiers() & Qt.ControlModifier:
            delta_y = event.angleDelta().y()
            if delta_y == 0:
                delta_y = event.pixelDelta().y()
            if delta_y > 0:
                self.zoom_in_step()
                event.accept()
                return
            elif delta_y < 0:
                self.zoom_out_step()
                event.accept()
                return
        super().wheelEvent(event)


class QCPreviewDialog(QDialog):
    windowClosed = pyqtSignal()
    pathSubmitted = pyqtSignal(str)

    def __init__(
        self,
        parent=None,
        folder_path: str = "",
        html_content: str = "",
        theme_name: str = "dark",
    ):
        super().__init__(parent)
        self.setWindowTitle("QC Preview")
        self.resize(1100, 760)
        self.setMinimumSize(820, 560)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        self.path_frame = QFrame()
        self.path_frame.setObjectName("sectionCard")
        self.path_layout = QVBoxLayout(self.path_frame)
        self.path_layout.setContentsMargins(12, 12, 12, 12)
        self.path_layout.setSpacing(6)

        self.path_title = QLabel("Folder Path")
        self.path_title.setObjectName("previewTitle")

        self.path_row = QHBoxLayout()
        self.path_row.setSpacing(8)

        self.path_value = QLineEdit()
        self.path_value.setPlaceholderText("Enter folder path here and press Enter...")
        self.path_value.returnPressed.connect(self._emit_path_submit)

        self.path_apply_btn = QPushButton("Apply")
        self.path_apply_btn.clicked.connect(self._emit_path_submit)
        self.path_apply_btn.setFixedHeight(34)

        self.zoom_in_btn = QPushButton("A+")
        self.zoom_in_btn.setToolTip("Zoom In")
        self.zoom_in_btn.setFixedHeight(34)
        self.zoom_in_btn.setFixedWidth(44)

        self.zoom_out_btn = QPushButton("A-")
        self.zoom_out_btn.setToolTip("Zoom Out")
        self.zoom_out_btn.setFixedHeight(34)
        self.zoom_out_btn.setFixedWidth(44)

        self.zoom_reset_btn = QPushButton("A")
        self.zoom_reset_btn.setToolTip("Reset Zoom")
        self.zoom_reset_btn.setFixedHeight(34)
        self.zoom_reset_btn.setFixedWidth(44)

        self.path_hint = QLabel(
            "Press Enter or Apply \u2022 Ctrl + Mouse Wheel / Ctrl + +/- / Zoom Buttons"
        )
        self.path_hint.setObjectName("hintLabel")

        self.path_row.addWidget(self.path_value, 1)
        self.path_row.addWidget(self.path_apply_btn)
        self.path_row.addWidget(self.zoom_in_btn)
        self.path_row.addWidget(self.zoom_out_btn)
        self.path_row.addWidget(self.zoom_reset_btn)

        self.path_layout.addWidget(self.path_title)
        self.path_layout.addLayout(self.path_row)
        self.path_layout.addWidget(self.path_hint)

        self.preview_frame = QFrame()
        self.preview_frame.setObjectName("sectionCard")
        self.preview_layout = QVBoxLayout(self.preview_frame)
        self.preview_layout.setContentsMargins(12, 12, 12, 12)
        self.preview_layout.setSpacing(8)

        self.preview_title = QLabel("Preview")
        self.preview_title.setObjectName("previewTitle")

        self.preview_box = ZoomableTextEdit()
        self.preview_box.setFocusPolicy(Qt.ClickFocus)
        self.preview_box.setReadOnly(True)
        self.preview_box.setAcceptRichText(True)
        self.preview_box.setLineWrapMode(QTextEdit.NoWrap)

        self.zoom_in_btn.clicked.connect(self._on_zoom_in_clicked)
        self.zoom_out_btn.clicked.connect(self._on_zoom_out_clicked)
        self.zoom_reset_btn.clicked.connect(self._on_zoom_reset_clicked)

        for key_seq, slot in [
            ("Ctrl++", self._on_zoom_in_clicked),
            ("Ctrl+=", self._on_zoom_in_clicked),
            ("Ctrl+-", self._on_zoom_out_clicked),
            ("Ctrl+0", self._on_zoom_reset_clicked),
        ]:
            shortcut = QShortcut(QKeySequence(key_seq), self)
            shortcut.activated.connect(slot)
            shortcut.setContext(Qt.WidgetWithChildrenShortcut)

        self.preview_layout.addWidget(self.preview_title)
        self.preview_layout.addWidget(self.preview_box, 1)

        layout.addWidget(self.path_frame)
        layout.addWidget(self.preview_frame, 1)

        self.installEventFilter(self)

        self.set_content(folder_path, html_content)
        self.apply_theme(theme_name)

        QTimer.singleShot(50, self.setFocus)
        QTimer.singleShot(100, lambda: self.preview_box.setFocus())

    def _on_zoom_in_clicked(self):
        self.preview_box.zoom_in_step()

    def _on_zoom_out_clicked(self):
        self.preview_box.zoom_out_step()

    def _on_zoom_reset_clicked(self):
        self.preview_box.reset_zoom()

    def eventFilter(self, obj, event):
        if event.type() == QEvent.Wheel:
            if event.modifiers() & Qt.ControlModifier:
                delta_y = event.angleDelta().y()
                if delta_y == 0:
                    delta_y = event.pixelDelta().y()
                if delta_y > 0:
                    self.preview_box.zoom_in_step()
                    event.accept()
                    return True
                elif delta_y < 0:
                    self.preview_box.zoom_out_step()
                    event.accept()
                    return True
        return super().eventFilter(obj, event)

    def _emit_path_submit(self):
        self.pathSubmitted.emit(self.path_value.text().strip())

    def set_content(self, folder_path: str, html_content: str):
        current_text = self.path_value.text().strip()
        incoming_text = (folder_path or "").strip()
        if not current_text or current_text == incoming_text:
            self.path_value.setText(incoming_text)
        self.preview_box.setHtml(html_content or "")

    def apply_theme(self, theme_name: str):
        theme = THEMES[theme_name]
        self.setStyleSheet(
            f"""
            QDialog {{
                background: {theme["root_bg"]};
            }}
            QFrame#sectionCard {{
                background: {theme["panel_bg"]};
                border: 1px solid {theme["border"]};
                border-radius: 14px;
            }}
            QLabel#previewTitle {{
                color: {theme["text_fg"]};
                font-size: 15px;
                font-weight: 800;
            }}
            QLabel#hintLabel {{
                color: {theme["muted_fg"]};
                font-size: 11px;
            }}
            QTextEdit, QLineEdit {{
                background: {theme["text_bg"]};
                color: {theme["text_fg"]};
                border: 1px solid {theme["border"]};
                border-radius: 10px;
                padding: 10px;
            }}
            QPushButton {{
                background: {theme["button_alt_bg"]};
                color: {theme["button_alt_fg"]};
                border: 1px solid {theme["border"]};
                border-radius: 10px;
                padding: 8px 14px;
                font-weight: 700;
            }}
            QPushButton:hover {{
                border: 1px solid {theme["button_bg"]};
            }}
            """
        )

    def closeEvent(self, event):
        self.windowClosed.emit()
        super().closeEvent(event)


# ------------------------------------------------------------
# Main Window
# ------------------------------------------------------------
class MailPrepWindow(QMainWindow):
    COPY_FEEDBACK_MS = 1400

    def __init__(self):
        super().__init__()

        self.generated_html = ""
        self.generated_plain_text = ""
        self.generated_clipboard_html = ""
        self.generated_preview_html = ""
        self.generated_qc_preview_html = ""
        self.qc_preview_mode = False
        self.qc_preview_window = None
        self.generated_copy_html = ""
        self.generated_copy_plain_text = ""
        self.generated_full_mail_html = ""
        self.generated_full_mail_plain_text = ""
        self.generated_notes_only_html = ""
        self.generated_notes_only_plain_text = ""

        self.last_preview_signature = None
        self.loading_active = False
        self.pending_preview_job: Optional[Tuple[str, object, bool]] = None
        self.current_job_id = 0
        self.current_theme = CURRENT_THEME
        self.last_refresh_display = "--"

        self.worker_thread: Optional[QThread] = None
        self.worker: Optional[PreviewWorker] = None

        self.shot_note_edits: Dict[str, QTextEdit] = {}
        self.shot_note_labels: Dict[str, QLabel] = {}
        self.shot_note_blocks: Dict[str, QFrame] = {}
        self.current_shot_names: List[str] = []
        self.current_raw_shot_names: List[str] = []
        self.current_duplicate_shot_names: List[str] = []
        self.current_result_lines: List[LineItem] = []
        self.current_suspicious_names: List[str] = []
        self.current_missing_summary_count = 0
        self.current_warning_messages: List[str] = []
        self.auto_shot_count_value: Optional[int] = None
        self.shot_count_auto_applied = False
        # Tracks if user has manually edited the shot count field
        self.shot_count_user_edited = False

        self.current_input_mode = "path"  # path | manifest
        self.current_manifest_data = None
        self.template_sheet_path = ""
        self.recent_paths: List[str] = []
        self.updating_recent_combo = False

        self.path_debounce_timer = QTimer(self)
        self.path_debounce_timer.setSingleShot(True)
        self.path_debounce_timer.timeout.connect(self._debounced_path_preview)

        # Debounce timer for metadata changes to avoid redundant re-renders while typing
        self.metadata_debounce_timer = QTimer(self)
        self.metadata_debounce_timer.setSingleShot(True)
        self.metadata_debounce_timer.timeout.connect(self._do_metadata_update)

        self.auto_refresh_timer = QTimer(self)
        self.auto_refresh_timer.timeout.connect(self.auto_refresh_tick)
        self.auto_refresh_timer.start(1000)

        self.copy_feedback_timer = QTimer(self)
        self.copy_feedback_timer.setSingleShot(True)
        self.copy_feedback_timer.timeout.connect(self._restore_copy_button_labels)

        self._build_ui()
        self._set_template_sheet_path("")
        self._setup_shortcuts()
        self._load_recent_paths()
        self._refresh_recent_paths_combo()
        self.apply_theme()
        self._rebuild_shot_notes_panel([])
        self._update_validation_summary()

    # ----------------------------------------------------------
    # UI Construction — broken into sub-methods for clarity
    # ----------------------------------------------------------
    def _build_ui(self):
        self.setWindowTitle("MailPrep Pro")
        self.resize(1460, 940)
        self.setMinimumSize(1180, 760)

        central = QWidget()
        self.setCentralWidget(central)

        self.outer_layout = QVBoxLayout(central)
        self.outer_layout.setContentsMargins(12, 12, 12, 12)
        self.outer_layout.setSpacing(10)

        self.outer_layout.addWidget(self._build_header())
        self.outer_layout.addWidget(self._build_control_card())
        self.outer_layout.addWidget(self._build_info_bar())

        self.main_content_row = QHBoxLayout()
        self.main_content_row.setSpacing(10)
        self.main_content_row.addWidget(self._build_preview_card(), 3)
        self.main_content_row.addWidget(self._build_notes_card(), 2)
        self.outer_layout.addLayout(self.main_content_row, 1)

        self.outer_layout.addWidget(self._build_status_bar())

    def _build_header(self) -> QFrame:
        self.header_frame = QFrame()
        self.header_frame.setObjectName("sectionCard")
        layout = QHBoxLayout(self.header_frame)
        layout.setContentsMargins(14, 12, 14, 12)
        layout.setSpacing(10)

        self.title_label = QLabel("\U0001f4e6 MailPrep Pro Tool")
        self.title_label.setObjectName("titleLabel")

        self.subtitle_label = QLabel("new way to future")
        self.subtitle_label.setObjectName("subtitleLabel")

        title_col = QVBoxLayout()
        title_col.setSpacing(2)
        title_col.addWidget(self.title_label)
        title_col.addWidget(self.subtitle_label)

        self.theme_btn = QPushButton("\u2600 Light Mode")
        self.theme_btn.clicked.connect(self.toggle_theme)

        layout.addLayout(title_col, 1)
        layout.addWidget(self.theme_btn, 0, Qt.AlignRight | Qt.AlignVCenter)
        return self.header_frame

    def _build_control_card(self) -> QFrame:
        self.control_card = QFrame()
        self.control_card.setObjectName("sectionCard")
        self.control_layout = QVBoxLayout(self.control_card)
        self.control_layout.setContentsMargins(14, 14, 14, 14)
        self.control_layout.setSpacing(12)

        self.control_layout.addLayout(self._build_path_row())
        self.control_layout.addLayout(self._build_meta_grid())
        self.control_layout.addLayout(self._build_action_row())

        self.advanced_frame = QFrame()
        self.advanced_frame.setObjectName("advancedFrame")
        adv_layout = QHBoxLayout(self.advanced_frame)
        adv_layout.setContentsMargins(10, 10, 10, 10)
        adv_layout.setSpacing(18)

        self.include_subject_check = QCheckBox("Include Package Name in Preview")
        self.include_subject_check.setChecked(True)
        self.include_subject_check.toggled.connect(self.on_metadata_change)

        self.include_shot_count_check = QCheckBox("Include Shot Count in Preview")
        self.include_shot_count_check.setChecked(True)
        self.include_shot_count_check.toggled.connect(self.on_metadata_change)

        self.include_submission_notes_check = QCheckBox("Include Submission Notes")
        self.include_submission_notes_check.setChecked(True)
        self.include_submission_notes_check.toggled.connect(self.on_metadata_change)

        self.include_meta_in_copy_check = QCheckBox("Include Metadata in Copy Output")
        self.include_meta_in_copy_check.setChecked(True)
        self.include_meta_in_copy_check.toggled.connect(self.on_metadata_change)

        self.complex_mode_check = QCheckBox("Complex Shot Mode")
        self.complex_mode_check.setChecked(False)
        self.complex_mode_check.toggled.connect(self.on_metadata_change)

        adv_layout.addWidget(self.include_subject_check)
        adv_layout.addWidget(self.include_shot_count_check)
        adv_layout.addWidget(self.include_submission_notes_check)
        adv_layout.addWidget(self.include_meta_in_copy_check)
        adv_layout.addWidget(self.complex_mode_check)
        adv_layout.addStretch(1)

        self.advanced_frame.hide()
        self.control_layout.addWidget(self.advanced_frame)

        return self.control_card

    def _build_path_row(self) -> QHBoxLayout:
        row = QHBoxLayout()
        row.setSpacing(8)

        self.path_label = QLabel("Folder")
        self.path_label.setObjectName("sectionLabel")

        self.path_edit = QLineEdit()
        self.path_edit.setPlaceholderText("Select or paste folder path...")
        self.path_edit.textChanged.connect(self.on_path_change)

        self.recent_paths_combo = QComboBox()
        self.recent_paths_combo.setObjectName("recentPathsCombo")
        self.recent_paths_combo.setMinimumWidth(220)
        self.recent_paths_combo.currentIndexChanged.connect(self.on_recent_path_selected)

        self.browse_btn = QPushButton("\U0001f4c1 Browse")
        self.browse_btn.clicked.connect(self.select_folder)

        self.warnings_btn = QPushButton("\u26a0 Warnings")
        self.warnings_btn.clicked.connect(self.show_warnings)

        self.generate_btn = QPushButton("\u2699 Generate Preview")
        self.generate_btn.setObjectName("primaryButton")
        self.generate_btn.clicked.connect(lambda: self.generate_preview(auto=False))
        self.generate_btn.setFixedHeight(34)

        self.auto_refresh_checkbox = QCheckBox("Auto Refresh")
        self.auto_refresh_checkbox.setChecked(True)
        self.auto_refresh_checkbox.toggled.connect(self.on_auto_refresh_toggle)

        row.addWidget(self.path_label)
        row.addWidget(self.path_edit, 1)
        row.addWidget(self.recent_paths_combo)
        row.addWidget(self.browse_btn)
        row.addWidget(self.warnings_btn)
        row.addWidget(self.generate_btn)
        row.addWidget(self.auto_refresh_checkbox)
        return row

    def _build_meta_grid(self) -> QGridLayout:
        grid = QGridLayout()
        grid.setHorizontalSpacing(10)
        grid.setVerticalSpacing(8)

        self.subject_label = QLabel("Package Name")
        self.subject_label.setObjectName("sectionLabel")

        self.subject_edit = QLineEdit()
        self.subject_edit.setPlaceholderText("Optional package name...")
        self.subject_edit.textChanged.connect(self.on_metadata_change)

        self.shot_count_label = QLabel("Shot Count")
        self.shot_count_label.setObjectName("sectionLabel")

        self.shot_count_edit = QLineEdit()
        self.shot_count_edit.setPlaceholderText("Optional")
        self.shot_count_edit.setFixedWidth(120)
        self.shot_count_edit.textChanged.connect(self.on_metadata_change)
        self.shot_count_edit.textEdited.connect(self.on_shot_count_edited)

        self.preview_dialog_btn = QPushButton("\U0001f5bc QC Preview: OFF")
        self.preview_dialog_btn.setCheckable(True)
        self.preview_dialog_btn.setChecked(False)
        self.preview_dialog_btn.clicked.connect(self.toggle_qc_preview_mode)

        self.clear_btn = QPushButton("\U0001f9f9 Clear / Reset")
        self.clear_btn.clicked.connect(self.clear_form)

        grid.addWidget(self.subject_label, 0, 0)
        grid.addWidget(self.subject_edit, 0, 1)
        grid.addWidget(self.shot_count_label, 0, 2)
        grid.addWidget(self.shot_count_edit, 0, 3)
        grid.addWidget(self.preview_dialog_btn, 0, 4)
        grid.addWidget(self.clear_btn, 0, 5)
        return grid

    def _build_action_row(self) -> QHBoxLayout:
        row = QHBoxLayout()
        row.setSpacing(8)

        self.copy_rich_btn = QPushButton("\U0001f4cb Copy Rich Text")
        self.copy_rich_btn.setObjectName("copyRichButton")
        self.copy_rich_btn.clicked.connect(self.copy_rich_text)

        self.copy_html_btn = QPushButton("</> Copy HTML")
        self.copy_html_btn.setObjectName("copyHtmlButton")
        self.copy_html_btn.clicked.connect(self.copy_html_source)

        self.copy_full_mail_btn = QPushButton("\U0001f4e8 Copy Full Mail")
        self.copy_full_mail_btn.setObjectName("copyFullButton")
        self.copy_full_mail_btn.clicked.connect(self.copy_full_mail)

        self.copy_full_mail_html_btn = QPushButton("\U0001f4e8 </> Copy Full Mail HTML")
        self.copy_full_mail_html_btn.setObjectName("copyFullButton")
        self.copy_full_mail_html_btn.clicked.connect(self.copy_full_mail_html)

        self.attach_sheet_btn = QPushButton("\U0001f4ce Attach Sheet")
        self.attach_sheet_btn.setObjectName("copyNotesButton")
        self.attach_sheet_btn.clicked.connect(self.attach_template_sheet)

        self.tools_btn = QToolButton()
        self.tools_btn.setText("\U0001f6e0 Tools")
        self.tools_btn.setPopupMode(QToolButton.InstantPopup)
        self.tools_menu = QMenu(self)
        self.tools_btn.setMenu(self.tools_menu)
        self._build_tools_menu()

        self.advanced_toggle_btn = QToolButton()
        self.advanced_toggle_btn.setText("\u25b6 Advanced Options")
        self.advanced_toggle_btn.setCheckable(True)
        self.advanced_toggle_btn.setChecked(False)
        self.advanced_toggle_btn.clicked.connect(self.toggle_advanced_options)

        row.addWidget(self.copy_rich_btn)
        row.addWidget(self.copy_html_btn)
        row.addWidget(self.copy_full_mail_btn)
        row.addWidget(self.copy_full_mail_html_btn)
        row.addWidget(self.attach_sheet_btn)
        row.addWidget(self.tools_btn)
        row.addStretch(1)
        row.addWidget(self.advanced_toggle_btn)
        return row

    def _build_info_bar(self) -> QFrame:
        self.info_frame = QFrame()
        self.info_frame.setObjectName("sectionCard")
        layout = QHBoxLayout(self.info_frame)
        layout.setContentsMargins(14, 10, 14, 10)
        layout.setSpacing(8)

        self.preview_count_label = QLabel("Folders: 0   |   Files: 0")
        self.preview_count_label.setObjectName("previewCountLabel")

        self.validation_label = QLabel("\u2705 Ready")
        self.validation_label.setObjectName("validationLabel")

        self.last_refresh_label = QLabel("Last Refresh: --")
        self.last_refresh_label.setObjectName("hintLabel")

        self.source_label = QLabel("Source: Folder Path")
        self.source_label.setObjectName("hintLabel")

        self.loading_label = QLabel("Scanning folder...")
        self.loading_label.setObjectName("loadingLabel")
        self.loading_label.hide()

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 0)
        self.progress_bar.setFixedWidth(140)
        self.progress_bar.hide()

        self.hint_label = QLabel("Auto-refresh is ON")
        self.hint_label.setObjectName("hintLabel")

        layout.addWidget(self.preview_count_label, 0, Qt.AlignLeft)
        layout.addWidget(self.validation_label, 0, Qt.AlignLeft)
        layout.addWidget(self.last_refresh_label, 0, Qt.AlignLeft)
        layout.addWidget(self.source_label, 0, Qt.AlignLeft)
        layout.addStretch(1)
        layout.addWidget(self.loading_label, 0, Qt.AlignRight)
        layout.addWidget(self.progress_bar, 0, Qt.AlignRight)
        layout.addWidget(self.hint_label, 0, Qt.AlignRight)
        return self.info_frame

    def _build_preview_card(self) -> QFrame:
        self.preview_card = QFrame()
        self.preview_card.setObjectName("previewCard")
        layout = QVBoxLayout(self.preview_card)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)

        self.preview_header = QFrame()
        self.preview_header.setObjectName("previewHeader")
        header_layout = QHBoxLayout(self.preview_header)
        header_layout.setContentsMargins(10, 8, 10, 8)

        self.preview_title = QLabel("Preview")
        self.preview_title.setObjectName("previewTitle")

        self.preview_subtitle = QLabel("One-shot rendered output")
        self.preview_subtitle.setObjectName("hintLabel")

        head_col = QVBoxLayout()
        head_col.setSpacing(0)
        head_col.addWidget(self.preview_title)
        head_col.addWidget(self.preview_subtitle)
        header_layout.addLayout(head_col, 1)

        self.preview_box = QTextEdit()
        self.preview_box.setReadOnly(True)
        self.preview_box.setAcceptRichText(True)
        self.preview_box.setLineWrapMode(QTextEdit.NoWrap)

        layout.addWidget(self.preview_header)
        layout.addWidget(self.preview_box, 1)
        return self.preview_card

    def _build_notes_card(self) -> QFrame:
        self.notes_card = QFrame()
        self.notes_card.setObjectName("sectionCard")
        layout = QVBoxLayout(self.notes_card)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)

        self.notes_title = QLabel("Shot Submission Notes")
        self.notes_title.setObjectName("previewTitle")

        self.notes_hint = QLabel("Editable notes for each derived version name")
        self.notes_hint.setObjectName("hintLabel")

        notes_head = QVBoxLayout()
        notes_head.setSpacing(0)
        notes_head.addWidget(self.notes_title)
        notes_head.addWidget(self.notes_hint)

        self.notes_search_edit = QLineEdit()
        self.notes_search_edit.setPlaceholderText("Search / filter shot notes...")
        self.notes_search_edit.textChanged.connect(self.filter_notes_panel)

        self.apply_all_row = QHBoxLayout()
        self.apply_all_row.setSpacing(8)

        self.common_note_edit = QLineEdit()
        self.common_note_edit.setPlaceholderText(
            "Common text to apply for all shots..."
        )
        self.common_note_edit.returnPressed.connect(self.apply_common_text_to_all_shots)

        self.apply_all_btn = QPushButton("\u21e2 Apply to All Shots")
        self.apply_all_btn.clicked.connect(self.apply_common_text_to_all_shots)

        self.apply_all_row.addWidget(self.common_note_edit, 1)
        self.apply_all_row.addWidget(self.apply_all_btn)

        self.notes_scroll = QScrollArea()
        self.notes_scroll.setWidgetResizable(True)
        self.notes_scroll.setFrameShape(QFrame.NoFrame)

        self.notes_scroll_widget = QWidget()
        self.notes_scroll_layout = QVBoxLayout(self.notes_scroll_widget)
        self.notes_scroll_layout.setContentsMargins(0, 0, 0, 0)
        self.notes_scroll_layout.setSpacing(10)

        self.notes_scroll.setWidget(self.notes_scroll_widget)

        layout.addLayout(notes_head)
        layout.addWidget(self.notes_search_edit)
        layout.addLayout(self.apply_all_row)
        layout.addWidget(self.notes_scroll, 1)
        return self.notes_card

    def _build_status_bar(self) -> QFrame:
        self.status_frame = QFrame()
        self.status_frame.setObjectName("sectionCard")
        layout = QHBoxLayout(self.status_frame)
        layout.setContentsMargins(14, 10, 14, 10)

        self.status_label = QLabel("Ready.")
        self.status_label.setObjectName("statusLabel")

        layout.addWidget(self.status_label)
        return self.status_frame

    # ----------------------------------------------------------
    # QC Preview window management
    # ----------------------------------------------------------
    def _show_qc_preview_window(self):
        if self.qc_preview_window is None:
            self.qc_preview_window = QCPreviewDialog(
                self,
                folder_path=self.path_edit.text().strip(),
                html_content=self.generated_qc_preview_html,
                theme_name=self.current_theme,
            )
            self.qc_preview_window.windowClosed.connect(
                self._on_qc_preview_window_closed
            )
            self.qc_preview_window.pathSubmitted.connect(
                self._on_qc_preview_path_submitted
            )
        else:
            self.qc_preview_window.apply_theme(self.current_theme)
            self.qc_preview_window.set_content(
                self.path_edit.text().strip(),
                self.generated_qc_preview_html,
            )

        self.qc_preview_window.show()
        self.qc_preview_window.raise_()
        self.qc_preview_window.activateWindow()
        self.qc_preview_window.exec_()

    def _on_qc_preview_path_submitted(self, path: str):
        path = (path or "").strip()
        if not path:
            self.set_status("Enter a valid folder path in QC Preview.")
            return

        self.current_input_mode = "path"
        self.current_manifest_data = None
        self.auto_refresh_checkbox.setEnabled(True)
        self.source_label.setText("Source: Folder Path")
        self.path_edit.setText(path)

        if not os.path.isdir(path):
            self.set_status("QC Preview path is invalid.")
            return

        self.start_preview_job(path, auto=False)

    def _close_qc_preview_window(self):
        if self.qc_preview_window is not None:
            self.qc_preview_window.close()

    def _on_qc_preview_window_closed(self):
        if self.preview_dialog_btn.isChecked():
            self.preview_dialog_btn.blockSignals(True)
            self.preview_dialog_btn.setChecked(False)
            self.preview_dialog_btn.blockSignals(False)
        self.qc_preview_mode = False
        self.preview_dialog_btn.setText("\U0001f5bc QC Preview: OFF")
        self.qc_preview_window = None
        self.set_status("QC preview mode disabled.")

    def _sync_qc_preview_window(self):
        if self.qc_preview_window is not None:
            self.qc_preview_window.apply_theme(self.current_theme)
            self.qc_preview_window.set_content(
                self.path_edit.text().strip(),
                self.generated_qc_preview_html,
            )

    # ----------------------------------------------------------
    # Tools menu
    # ----------------------------------------------------------
    def _build_tools_menu(self):
        self.tools_menu.clear()

        import_notes_action = QAction("Import Notes", self)
        import_notes_action.triggered.connect(self.import_notes_file)

        export_notes_action = QAction("Export Notes", self)
        export_notes_action.triggered.connect(self.export_notes_file)

        export_csv_action = QAction("Export Submission CSV", self)
        export_csv_action.triggered.connect(self.export_submission_csv)

        self.tools_menu.addAction(import_notes_action)
        self.tools_menu.addAction(export_notes_action)
        self.tools_menu.addAction(export_csv_action)
        self.tools_menu.addSeparator()

        save_session_action = QAction("Save Session", self)
        save_session_action.triggered.connect(self.save_session_file)

        load_session_action = QAction("Load Session", self)
        load_session_action.triggered.connect(self.load_session_file)

        self.tools_menu.addAction(save_session_action)
        self.tools_menu.addAction(load_session_action)
        self.tools_menu.addSeparator()

        export_mail_html_action = QAction("Export Mail as HTML", self)
        export_mail_html_action.triggered.connect(self.export_mail_html)

        export_mail_txt_action = QAction("Export Mail as TXT", self)
        export_mail_txt_action.triggered.connect(self.export_mail_txt)

        self.tools_menu.addAction(export_mail_html_action)
        self.tools_menu.addAction(export_mail_txt_action)

    # ----------------------------------------------------------
    # Shortcuts
    # ----------------------------------------------------------
    def _setup_shortcuts(self):
        QShortcut(QKeySequence("Ctrl+B"), self, activated=self.select_folder)
        QShortcut(
            QKeySequence("Ctrl+R"),
            self,
            activated=lambda: self.generate_preview(auto=False),
        )
        QShortcut(QKeySequence("Ctrl+Shift+R"), self, activated=self.copy_rich_text)
        QShortcut(QKeySequence("Ctrl+Shift+H"), self, activated=self.copy_html_source)
        QShortcut(QKeySequence("Ctrl+Shift+M"), self, activated=self.copy_full_mail)
        QShortcut(QKeySequence("Ctrl+Shift+N"), self, activated=self.copy_notes_only)
        QShortcut(QKeySequence("Ctrl+K"), self, activated=self.clear_form)
        QShortcut(QKeySequence("Ctrl+T"), self, activated=self.toggle_theme)
        QShortcut(QKeySequence("Ctrl+P"), self, activated=self.open_preview_dialog)
        QShortcut(QKeySequence("Ctrl+M"), self, activated=self.load_manifest_file)
        QShortcut(QKeySequence("Ctrl+Shift+S"), self, activated=self.save_session_file)
        QShortcut(QKeySequence("Ctrl+Shift+L"), self, activated=self.load_session_file)

    # ----------------------------------------------------------
    # Theming
    # ----------------------------------------------------------
    def apply_theme(self):
        theme = THEMES[self.current_theme]

        self.setStyleSheet(
            f"""
            QMainWindow {{
                background: {theme["root_bg"]};
            }}
            QFrame {{
                background: {theme["panel_bg"]};
                border: none;
            }}
            QFrame#sectionCard {{
                background: {theme["panel_bg"]};
                border: 1px solid {theme["border"]};
                border-radius: 14px;
            }}
            QFrame#advancedFrame {{
                background: {theme["toolbar_bg"]};
                border: 1px solid {theme["border"]};
                border-radius: 10px;
            }}
            #titleLabel {{
                color: {theme["text_fg"]};
                font-size: 20px;
                font-weight: 800;
                background: transparent;
            }}
            #subtitleLabel {{
                color: {theme["muted_fg"]};
                font-size: 12px;
                background: transparent;
            }}
            #sectionLabel {{
                color: {theme["section_title_fg"]};
                font-size: 12px;
                font-weight: 700;
                background: transparent;
                padding-bottom: 2px;
            }}
            QLineEdit, QTextEdit, QComboBox {{
                background: {theme["entry_bg"]};
                color: {theme["entry_fg"]};
                border: 1px solid {theme["border"]};
                border-radius: 12px;
                padding: 8px 10px;
                selection-background-color: {theme["button_bg"]};
            }}
            QComboBox::drop-down {{
                border: none;
                width: 22px;
            }}
            QComboBox QAbstractItemView {{
                background: {theme["entry_bg"]};
                color: {theme["entry_fg"]};
                border: 1px solid {theme["border"]};
                selection-background-color: {theme["button_bg"]};
            }}
            QPushButton {{
                background: {theme["button_alt_bg"]};
                color: {theme["button_alt_fg"]};
                border: 1px solid {theme["border"]};
                border-radius: 12px;
                padding: 9px 13px;
                font-weight: 700;
            }}
            QPushButton:hover {{
                border: 1px solid {theme["button_bg"]};
            }}
            QPushButton#primaryButton {{
                background: {theme["primary_button_bg"]};
                color: {theme["button_fg"]};
                border: none;
                border-radius: 16px;
                padding: 6px 16px;
                font-size: 12px;
                font-weight: 800;
                min-width: 130px;
            }}
            QPushButton#primaryButton:hover {{
                background: {theme["primary_button_hover"]};
            }}
            QPushButton#copyRichButton {{
                background: {theme["copy_rich_bg"]};
                color: {theme["copy_rich_fg"]};
                border: 1px solid {theme["copy_rich_border"]};
                border-radius: 12px;
                font-weight: 800;
            }}
            QPushButton#copyHtmlButton {{
                background: {theme["copy_html_bg"]};
                color: {theme["copy_html_fg"]};
                border: 1px solid {theme["copy_html_border"]};
                border-radius: 12px;
                font-weight: 800;
            }}
            QPushButton#copyFullButton {{
                background: {theme["copy_full_bg"]};
                color: {theme["copy_full_fg"]};
                border: 1px solid {theme["copy_full_border"]};
                border-radius: 12px;
                font-weight: 800;
            }}
            QPushButton#copyNotesButton {{
                background: {theme["copy_notes_bg"]};
                color: {theme["copy_notes_fg"]};
                border: 1px solid {theme["copy_notes_border"]};
                border-radius: 12px;
                font-weight: 800;
            }}
            QToolButton {{
                background: transparent;
                color: {theme["text_fg"]};
                border: 1px solid transparent;
                border-radius: 10px;
                font-weight: 700;
                padding: 6px 10px;
            }}
            QToolButton:hover {{
                border: 1px solid {theme["border"]};
                background: {theme["toolbar_bg"]};
            }}
            QMenu {{
                background: {theme["panel_bg"]};
                color: {theme["text_fg"]};
                border: 1px solid {theme["border"]};
            }}
            QMenu::item:selected {{
                background: {theme["toolbar_bg"]};
            }}
            QLabel {{
                color: {theme["text_fg"]};
                background: transparent;
            }}
            QCheckBox {{
                color: {theme["text_fg"]};
                spacing: 6px;
                background: transparent;
            }}
            #hintLabel, #loadingLabel, #statusLabel {{
                color: {theme["muted_fg"]};
            }}
            #previewCountLabel {{
                background: {theme["preview_count_bg"]};
                color: {theme["preview_count_fg"]};
                border-radius: 10px;
                padding: 6px 10px;
                font-weight: 800;
            }}
            #validationLabel {{
                border-radius: 10px;
                padding: 6px 10px;
                font-weight: 800;
            }}
            #previewCard {{
                background: {theme["panel_bg"]};
                border-radius: 14px;
            }}
            #previewHeader {{
                background: {theme["preview_header_bg"]};
                border: 1px solid {theme["preview_header_border"]};
                border-radius: 12px;
            }}
            #previewTitle {{
                color: {theme["text_fg"]};
                font-size: 15px;
                font-weight: 800;
            }}
            QProgressBar {{
                border: 1px solid {theme["border"]};
                border-radius: 8px;
                background: {theme["panel_bg"]};
                text-align: center;
            }}
            QProgressBar::chunk {{
                background: {theme["progress_bg"]};
                border-radius: 8px;
            }}
            """
        )

        self.theme_btn.setText(
            "\u2600 Light Mode" if self.current_theme == "dark" else "\U0001f319 Dark Mode"
        )

        self._apply_validation_style()
        self._refresh_note_label_styles()

        # Re-render from cached lines — avoids re-scanning the folder on theme change
        if self.current_result_lines:
            try:
                result = MailPrepLogic.build_result_from_lines(
                    self.current_result_lines,
                    self.current_theme,
                    self.last_preview_signature,
                )
                self._render_html_only(result)
            except Exception:
                pass

    def toggle_theme(self):
        self.current_theme = "light" if self.current_theme == "dark" else "dark"
        self.apply_theme()
        self.set_status(f"Theme changed to {self.current_theme} mode.")

    def _render_html_only(self, result: BuildResult):
        """Re-renders HTML outputs from a BuildResult without touching shot names or notes."""
        self.generated_html = result.html_text
        self.generated_clipboard_html = result.clipboard_html
        self.generated_preview_html = self._compose_preview_html(result.html_text)
        self.generated_qc_preview_html = self._compose_qc_preview_html(result.html_text)
        self.generated_copy_html = self._compose_copy_rich_html(result.clipboard_html)
        self.generated_copy_plain_text = self._compose_copy_plain_text(result.plain_text)
        self.generated_notes_only_html = self._compose_notes_only_html()
        self.generated_notes_only_plain_text = self._compose_notes_only_plain_text()
        self.generated_full_mail_html = self._compose_full_mail_html()
        self.generated_full_mail_plain_text = self._compose_full_mail_plain_text()
        self.preview_box.setHtml(self.generated_preview_html)
        self._sync_qc_preview_window()

    # ----------------------------------------------------------
    # Status / loading helpers
    # ----------------------------------------------------------
    def set_status(self, message: str):
        self.status_label.setText(message)

    def set_loading(self, is_loading: bool, message: str = "Loading..."):
        self.loading_active = is_loading
        self.loading_label.setText(message)
        self.loading_label.setVisible(is_loading)
        self.progress_bar.setVisible(is_loading)
        for btn in (self.generate_btn, self.browse_btn, self.attach_sheet_btn, self.warnings_btn):
            btn.setDisabled(is_loading)

    def _update_last_refresh(self):
        self.last_refresh_display = datetime.now().strftime("%d-%m-%Y %I:%M:%S %p")
        self.last_refresh_label.setText(f"Last Refresh: {self.last_refresh_display}")

    def _restore_copy_button_labels(self):
        self.copy_rich_btn.setText("\U0001f4cb Copy Rich Text")
        self.copy_html_btn.setText("</> Copy HTML")
        self.copy_full_mail_btn.setText("\U0001f4e8 Copy Full Mail")
        self.copy_full_mail_html_btn.setText("\U0001f4e8 </> Copy Full Mail HTML")

    # ----------------------------------------------------------
    # HTML composition helpers
    # ----------------------------------------------------------
    def _escape_with_breaks(self, value: str) -> str:
        return html.escape(value).replace("\n", "<br>")

    def _copy_safe_div(
        self,
        text: str,
        bold: bool = False,
        underline: bool = False,
        suspicious: bool = False,
    ) -> str:
        safe_text = html.escape(text).replace("\n", "<br>")
        if underline:
            safe_text = f"<u>{safe_text}</u>"
        if bold:
            safe_text = f"<b>{safe_text}</b>"
        return safe_text

    def _looks_like_file_name(self, name: str) -> bool:
        return bool(re.search(r"\.[A-Za-z0-9]{2,8}$", name.strip()))

    def _is_ignored_submission_parent(self, name: str) -> bool:
        n = name.strip().lower()

        if n in IGNORED_SUBMISSION_EXACT:
            return True

        tokens = [token for token in re.split(r"[._\-\s]+", n) if token]
        if not tokens:
            return False

        if all(
            token in IGNORED_SUBMISSION_EXACT or token in IGNORED_SUBMISSION_FUZZY
            for token in tokens
        ):
            return True

        return False

    def _cleanup_version_candidate(self, name: str) -> str:
        if not name:
            return ""

        candidate = name.strip()
        root, ext = os.path.splitext(candidate)
        if ext and ext.lower() in KNOWN_FILE_EXTENSIONS:
            candidate = root

        candidate = re.sub(r"\.\d+-\d+$", "", candidate)
        candidate = re.sub(r"\.\d+$", "", candidate)

        changed = True
        while changed:
            changed = False
            for suffix in REMOVABLE_VERSION_SUFFIXES:
                pattern = re.compile(rf"([_.-]){re.escape(suffix)}$", re.IGNORECASE)
                if pattern.search(candidate):
                    candidate = pattern.sub("", candidate)
                    changed = True

        candidate = candidate.strip("._- ")
        return candidate

    def _derive_submission_shot_name_from_filename(
        self, filename: str
    ) -> Optional[str]:
        if not filename:
            return None
        candidate = self._cleanup_version_candidate(filename)
        if not candidate:
            return None
        if self._is_ignored_submission_parent(candidate):
            return None
        return candidate

    def _looks_like_version_name(self, name: str) -> bool:
        if not name:
            return False
        cleaned = self._cleanup_version_candidate(name)
        if not cleaned:
            return False
        if self._is_ignored_submission_parent(cleaned):
            return False
        return bool(
            re.search(r"(?:^|[_.-])(?:v|z)\d{1,4}(?=$|[_.-])", cleaned, re.IGNORECASE)
        )

    def _collect_shot_names_default_raw(self, lines: List[LineItem]) -> List[str]:
        results = []

        for item in lines:
            if item.item_type != "folder":
                continue
            raw = item.text.strip()
            if not raw:
                continue
            if self._is_ignored_submission_parent(raw):
                continue
            if not self._looks_like_version_name(raw):
                continue
            version_name = self._cleanup_version_candidate(raw)
            if version_name:
                results.append(version_name)

        if results:
            return results

        for item in lines:
            if item.item_type != "file":
                continue
            raw = item.text.strip()
            if not raw:
                continue
            if not self._looks_like_file_name(raw):
                continue
            version_name = self._derive_submission_shot_name_from_filename(raw)
            if not version_name:
                continue
            if not self._looks_like_version_name(version_name):
                continue
            results.append(version_name)

        return results

    def _candidate_priority_from_filename(self, filename: str) -> int:
        low = filename.lower()
        if low.endswith(".sfx"):
            return 0
        if low.endswith(".mb"):
            return 1
        if low.endswith(".ma"):
            return 2
        if low.endswith(".fbx"):
            return 3
        if low.endswith(".nk"):
            return 4
        return 99

    def _is_preview_like_candidate(self, candidate: str) -> bool:
        low = candidate.lower()
        preview_tokens = (
            "flattened",
            "grey",
            "gray",
            "wireframe",
            "rotolines",
            "overlay",
            "holdout",
            "perspective",
            "perspectivestab",
            "pointblast",
            "pointblastdigorychest",
            "shaded",
            "curves",
            "check",
            "rotocheck",
            "trkholdout",
            "trkperspective",
            "trkperspective2",
            "trkshaded",
            "trkwire",
            "trkcurves",
            "trkpointblast",
        )
        return any(token in low for token in preview_tokens)

    def _candidate_token_count(self, candidate: str) -> int:
        parts = [p for p in re.split(r"[._-]+", candidate) if p]
        return len(parts)

    def _collect_shot_names_complex_raw(self, lines: List[LineItem]) -> List[str]:
        results = []

        for item in lines:
            if item.item_type != "folder":
                continue
            if item.level != 0:
                continue
            raw = item.text.strip()
            if not raw:
                continue
            if self._is_ignored_submission_parent(raw):
                continue
            if not self._looks_like_version_name(raw):
                continue
            version_name = self._cleanup_version_candidate(raw)
            if version_name:
                results.append(version_name)

        if results:
            return results

        candidates_by_root: Dict[str, Dict[str, Tuple[int, int, int, int]]] = {}
        current_root = None
        file_index = 0

        for item in lines:
            if item.item_type == "folder" and item.level == 0:
                current_root = item.text.strip()
                if current_root and current_root not in candidates_by_root:
                    candidates_by_root[current_root] = {}
                continue

            if item.item_type != "file":
                continue

            raw = item.text.strip()
            if not raw:
                continue
            if not self._looks_like_file_name(raw):
                continue

            candidate = self._derive_submission_shot_name_from_filename(raw)
            if not candidate:
                continue
            if not self._looks_like_version_name(candidate):
                continue
            if self._is_preview_like_candidate(candidate):
                continue

            root_key = current_root or "__ungrouped__"
            if root_key not in candidates_by_root:
                candidates_by_root[root_key] = {}

            priority = self._candidate_priority_from_filename(raw)
            token_count = self._candidate_token_count(candidate)
            length = len(candidate)

            root_map = candidates_by_root[root_key]
            current_metrics = (priority, token_count, length, file_index)

            existing = root_map.get(candidate)
            if existing is None or current_metrics < existing:
                root_map[candidate] = current_metrics

            file_index += 1

        for root_key in list(candidates_by_root.keys()):
            root_map = candidates_by_root.get(root_key, {})
            if not root_map:
                continue
            best_candidate = sorted(
                root_map.items(),
                key=lambda kv: (kv[1][0], kv[1][1], kv[1][2], kv[1][3], kv[0].lower()),
            )[0][0]
            results.append(best_candidate)

        if results:
            return results

        return self._collect_shot_names_default_raw(lines)

    def _dedupe_preserve_order(self, values: List[str]) -> List[str]:
        seen = set()
        result = []
        for value in values:
            if value not in seen:
                seen.add(value)
                result.append(value)
        return result

    def _extract_shot_names_from_lines(self, lines: List[LineItem]) -> List[str]:
        if self.complex_mode_check.isChecked():
            raw = self._collect_shot_names_complex_raw(lines)
        else:
            raw = self._collect_shot_names_default_raw(lines)
        self.current_raw_shot_names = raw[:]
        return self._dedupe_preserve_order(raw)

    def _duplicate_shot_names_from_raw(self, raw_names: List[str]) -> List[str]:
        counts: Dict[str, int] = {}
        duplicates = []
        for name in raw_names:
            counts[name] = counts.get(name, 0) + 1
        for name in raw_names:
            if counts.get(name, 0) > 1 and name not in duplicates:
                duplicates.append(name)
        return duplicates

    def get_submission_notes_map(self) -> Dict[str, str]:
        data = {}
        for shot_name in self.current_shot_names:
            editor = self.shot_note_edits.get(shot_name)
            if not editor:
                continue
            data[shot_name] = editor.toPlainText()
        return data

    def _note_label_html(self, shot_name: str) -> str:
        theme = THEMES[self.current_theme]
        suspicious = MailPrepLogic.is_suspicious_name(shot_name)
        duplicate = shot_name in self.current_duplicate_shot_names

        safe_text = html.escape(shot_name)

        if suspicious:
            safe_text = (
                f'<span style="background:{theme["suspicious_bg"]}; '
                f'color:{theme["suspicious_fg"]}; '
                f'border:1px solid {theme["suspicious_border"]}; '
                'border-radius:4px; padding:1px 4px;">'
                f"{safe_text}</span>"
            )

        prefix = ""
        if duplicate:
            prefix = (
                f'<span style="color:{theme["duplicate_fg"]}; '
                'font-weight:800;">\u26a0 Duplicate </span>'
            )

        return f"{prefix}<span style='font-weight:800;'>{safe_text}</span>"

    def _note_block_style(self, shot_name: str) -> str:
        theme = THEMES[self.current_theme]
        border = theme["border"]
        bg = theme["panel_bg"]

        if shot_name in self.current_duplicate_shot_names:
            border = theme["duplicate_border"]
        elif MailPrepLogic.is_suspicious_name(shot_name):
            border = theme["suspicious_border"]

        return (
            f"QFrame {{ background: {bg}; border: 1px solid {border}; "
            "border-radius: 12px; }}"
        )

    def _refresh_note_label_styles(self):
        for shot_name, label in self.shot_note_labels.items():
            label.setTextFormat(Qt.RichText)
            label.setText(self._note_label_html(shot_name))
        for shot_name, block in self.shot_note_blocks.items():
            block.setStyleSheet(self._note_block_style(shot_name))

    def _rebuild_shot_notes_panel(self, shot_names: List[str]):
        old_values = self.get_submission_notes_map()

        while self.notes_scroll_layout.count():
            item = self.notes_scroll_layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

        self.shot_note_edits = {}
        self.shot_note_labels = {}
        self.shot_note_blocks = {}
        self.current_shot_names = shot_names[:]

        if not shot_names:
            empty_label = QLabel("Generate preview to load shot-wise notes fields.")
            empty_label.setObjectName("hintLabel")
            self.notes_scroll_layout.addWidget(empty_label)
            self.notes_scroll_layout.addStretch(1)
            return

        for shot_name in shot_names:
            block = QFrame()
            block.setObjectName("sectionCard")
            block.setStyleSheet(self._note_block_style(shot_name))

            block_layout = QVBoxLayout(block)
            block_layout.setContentsMargins(10, 10, 10, 10)
            block_layout.setSpacing(6)

            label = QLabel()
            label.setTextFormat(Qt.RichText)
            label.setText(self._note_label_html(shot_name))

            editor = QTextEdit()
            editor.setAcceptRichText(False)
            editor.setPlaceholderText("Paste shot-specific submission note here...")
            editor.setFixedHeight(78)
            editor.textChanged.connect(self.on_metadata_change)

            if shot_name in old_values:
                editor.setPlainText(old_values[shot_name])

            block_layout.addWidget(label)
            block_layout.addWidget(editor)

            self.shot_note_edits[shot_name] = editor
            self.shot_note_labels[shot_name] = label
            self.shot_note_blocks[shot_name] = block
            self.notes_scroll_layout.addWidget(block)

        self.notes_scroll_layout.addStretch(1)
        self.filter_notes_panel()
        self._refresh_note_label_styles()

    def apply_common_text_to_all_shots(self):
        common_text = self.common_note_edit.text().strip()
        if not common_text:
            self.set_status("Enter common text first.")
            return

        updated = 0
        for shot_name in self.current_shot_names:
            block = self.shot_note_blocks.get(shot_name)
            if block is not None and not block.isVisible():
                continue
            editor = self.shot_note_edits.get(shot_name)
            if editor is None:
                continue
            editor.setPlainText(common_text)
            updated += 1

        if updated:
            self.on_metadata_change()
            self.set_status(f"Applied common text to {updated} shot(s).")

    def filter_notes_panel(self):
        query = self.notes_search_edit.text().strip().lower()
        visible_count = 0

        for shot_name, block in self.shot_note_blocks.items():
            note_text = ""
            editor = self.shot_note_edits.get(shot_name)
            if editor is not None:
                note_text = editor.toPlainText().lower()

            matches = not query or query in shot_name.lower() or query in note_text
            block.setVisible(matches)
            if matches:
                visible_count += 1

        if self.current_shot_names:
            self.notes_hint.setText(
                f"Editable notes for each derived version name \u2022 Showing "
                f"{visible_count}/{len(self.current_shot_names)}"
            )
        else:
            self.notes_hint.setText("Editable notes for each derived version name")

    def _resolved_shot_count(self) -> str:
        manual = self.shot_count_edit.text().strip()
        if manual:
            return manual
        if self.auto_shot_count_value is not None and self.auto_shot_count_value > 0:
            return str(self.auto_shot_count_value)
        return ""

    def _meta_rows(self) -> List[Tuple[str, str]]:
        rows = []
        package_name = self.subject_edit.text().strip()
        shot_count = self._resolved_shot_count()

        if self.include_subject_check.isChecked() and package_name:
            rows.append(("Package Name", package_name))
        if self.include_shot_count_check.isChecked() and shot_count:
            rows.append(("Shot Count", shot_count))
        return rows

    def _copy_meta_rows(self) -> List[Tuple[str, str]]:
        if not self.include_meta_in_copy_check.isChecked():
            return []
        return self._meta_rows()

    def _submission_notes_items(self) -> List[Tuple[str, str]]:
        if not self.include_submission_notes_check.isChecked():
            return []
        items = []
        for shot_name in self.current_shot_names:
            editor = self.shot_note_edits.get(shot_name)
            if not editor:
                continue
            note = editor.toPlainText().strip()
            if note:
                items.append((shot_name, note))
        return items

    def _strip_outer_html(self, value: str) -> str:
        body_only = re.sub(r"^\s*<html><body>", "", value, flags=re.IGNORECASE)
        body_only = re.sub(r"</body></html>\s*$", "", body_only, flags=re.IGNORECASE)
        return body_only

    def _strip_outer_div(self, value: str) -> str:
        body_only = self._strip_outer_html(value)
        body_only = re.sub(r"^\s*<div[^>]*>", "", body_only, flags=re.IGNORECASE)
        body_only = re.sub(r"</div>\s*$", "", body_only, flags=re.IGNORECASE)
        return body_only

    def _compose_notes_section_preview_html(self) -> str:
        notes_items = self._submission_notes_items()
        if not notes_items:
            return ""

        parts = [
            '<div style="height:1em;"></div>',
            "<div><b><u>Submission Notes :</u></b></div>",
        ]

        for idx, (shot_name, note) in enumerate(notes_items):
            if MailPrepLogic.is_suspicious_name(shot_name):
                theme = THEMES[self.current_theme]
                shot_html = (
                    f'<span style="background:{theme["suspicious_bg"]}; '
                    f'color:{theme["suspicious_fg"]}; '
                    f'border:1px solid {theme["suspicious_border"]}; '
                    'border-radius:4px; padding:1px 4px;">'
                    f"{html.escape(shot_name)}</span>"
                )
                parts.append(f"<div><b>{shot_html}</b></div>")
            else:
                parts.append(f"<div><b>{html.escape(shot_name)}</b></div>")

            for line in note.splitlines():
                if line.strip():
                    parts.append(f"<div>{html.escape(line)}</div>")
                else:
                    parts.append("<div><br></div>")

            if idx < len(notes_items) - 1:
                parts.append("<div><br></div>")

        return "".join(str(x) for x in parts)

    def _compose_notes_section_clipboard_html(self) -> str:
        notes_items = self._submission_notes_items()
        if not notes_items:
            return ""

        parts = []
        parts.append(
            self._copy_safe_div("Submission Notes :", bold=True, underline=True)
        )
        parts.append("<br>")

        for idx, (shot_name, note) in enumerate(notes_items):
            parts.append(self._copy_safe_div(shot_name, bold=True))
            parts.append("<br>")

            note_lines = note.splitlines()
            if note_lines:
                for line in note_lines:
                    if line.strip():
                        parts.append(self._copy_safe_div(line))
                    parts.append("<br>")
            else:
                parts.append("<br>")

            if idx < len(notes_items) - 1:
                parts.append("<br>")

        return "".join(str(x) for x in parts)

    def _compose_notes_section_plain_text(self) -> str:
        notes_items = self._submission_notes_items()
        if not notes_items:
            return ""

        notes_block = ["Submission Notes :", ""]
        for idx, (shot_name, note) in enumerate(notes_items):
            notes_block.append(shot_name)
            notes_block.append(note)
            if idx < len(notes_items) - 1:
                notes_block.append("")
        return "\n".join(str(x) for x in notes_block)

    def _compose_copy_body_html(self, rows: List[Tuple[str, str]]) -> str:
        parts = []

        if rows:
            for label, value in rows:
                safe_value = html.escape(value).replace("\n", "<br>")
                parts.append(f"<b>{html.escape(label)}:</b> {safe_value}<br>")
            parts.append("<br>")

        tree_body = self._strip_outer_div(self.generated_clipboard_html)
        parts.append(tree_body)

        notes_html = self._compose_notes_section_clipboard_html()
        if notes_html:
            parts.append("<br>")
            parts.append(notes_html)

        return "".join(str(x) for x in parts)

    def _compose_copy_body_plain_text(self, rows: List[Tuple[str, str]]) -> str:
        blocks = []
        if rows:
            for label, value in rows:
                blocks.append(f"{label}: {value}")
        blocks.append(self.generated_plain_text)
        notes_text = self._compose_notes_section_plain_text()
        if notes_text:
            blocks.append(notes_text)
        return "\n\n".join(str(b) for b in blocks if b not in (None, ""))

    def _compose_preview_html(self, tree_html: str) -> str:
        rows = self._meta_rows()

        parts = [
            "<html><body>",
            (
                f'<div style="font-family:Calibri, Arial, sans-serif; '
                f'font-size:12pt; line-height:1.2; color:{THEMES[self.current_theme]["text_fg"]};">'
            ),
        ]

        if rows:
            theme = THEMES[self.current_theme]
            parts.append(
                f'<div style="margin-bottom:10px; padding:10px 12px; '
                f'border:1px solid {theme["preview_header_border"]}; '
                f'background:{theme["preview_header_bg"]}; border-radius:8px;">'
            )
            for label, value in rows:
                parts.append(
                    f'<div style="margin-bottom:4px;"><b>{html.escape(label)}:</b> '
                    f"{self._escape_with_breaks(value)}</div>"
                )
            parts.append("</div>")

        parts.append(self._strip_outer_html(tree_html))
        parts.append(self._compose_notes_section_preview_html())
        parts.append("</div></body></html>")
        return "".join(str(x) for x in parts)

    def _compose_qc_preview_html(self, tree_html: str) -> str:
        path_value = self.path_edit.text().strip()

        parts = [
            "<html><body>",
            (
                f'<div style="font-family:Calibri, Arial, sans-serif; '
                f'font-size:12pt; line-height:1.2; color:{THEMES[self.current_theme]["text_fg"]};">'
            ),
        ]

        if path_value:
            theme = THEMES[self.current_theme]
            parts.append(
                f'<div style="margin-bottom:10px; padding:10px 12px; '
                f'border:1px solid {theme["preview_header_border"]}; '
                f'background:{theme["preview_header_bg"]}; border-radius:8px;">'
                f'<div><b>Folder Path:</b> {html.escape(path_value)}</div>'
                "</div>"
            )

        parts.append(self._strip_outer_html(tree_html))
        parts.append("</div></body></html>")
        return "".join(str(x) for x in parts)

    def _compose_copy_rich_html(self, tree_clipboard_html: str) -> str:
        rows = self._copy_meta_rows()
        inner = self._compose_copy_body_html(rows)
        return (
            "<html><body>"
            '<div style="font-family:Calibri, Arial, sans-serif; '
            'font-size:12pt; line-height:1.2; color:#000000; background-color:transparent;">'
            f"{inner}"
            "</div></body></html>"
        )

    def _compose_copy_plain_text(self, tree_plain_text: str) -> str:
        _ = tree_plain_text
        rows = self._copy_meta_rows()
        return self._compose_copy_body_plain_text(rows)

    def _compose_notes_only_html(self) -> str:
        notes_html = self._compose_notes_section_clipboard_html()
        if not notes_html:
            return ""
        return (
            "<html><body>"
            '<div style="font-family:Calibri, Arial, sans-serif; '
            'font-size:12pt; line-height:1.2; color:#000000; background-color:transparent;">'
            f"{notes_html}"
            "</div></body></html>"
        )

    def _compose_notes_only_plain_text(self) -> str:
        return self._compose_notes_section_plain_text()

    def _compose_full_mail_html(self) -> str:
        if not self.generated_plain_text:
            return ""
        inner = self._compose_copy_body_html(self._meta_rows())
        parts = [
            "<html><body>",
            '<div style="font-family:Calibri, Arial, sans-serif; '
            'font-size:12pt; line-height:1.2; color:#000000; background-color:transparent;">',
            "Hi Team,<br><br>",
            "Please find below the submission details.<br><br>",
            inner,
            "<br><br>Thanks.",
            "</div></body></html>",
        ]
        return "".join(str(x) for x in parts)

    def _compose_full_mail_plain_text(self) -> str:
        if not self.generated_plain_text:
            return ""
        body_text = self._compose_copy_body_plain_text(self._meta_rows())
        return (
            "Hi Team,\n\n"
            "Please find below the submission details.\n\n"
            f"{body_text}\n\n"
            "Thanks."
        )

    # ----------------------------------------------------------
    # Clipboard
    # ----------------------------------------------------------
    def _copy_html_to_clipboard(self, rich_html: str, plain_text: str):
        mime = QMimeData()
        mime.setHtml(rich_html)
        mime.setText(plain_text)
        mime.setData("text/html", rich_html.encode("utf-8"))

        clipboard = QGuiApplication.clipboard()
        clipboard.setMimeData(mime, QClipboard.Clipboard)

        try:
            mime2 = QMimeData()
            mime2.setHtml(rich_html)
            mime2.setText(plain_text)
            mime2.setData("text/html", rich_html.encode("utf-8"))
            clipboard.setMimeData(mime2, QClipboard.Selection)
        except Exception:
            pass

    def _set_copy_success(self, button: QPushButton, text: str, status_text: str):
        button.setText(text)
        self.copy_feedback_timer.start(self.COPY_FEEDBACK_MS)
        self.set_status(status_text)

    # ----------------------------------------------------------
    # Validation
    # ----------------------------------------------------------
    def _collect_suspicious_names(self) -> List[str]:
        names = []
        seen = set()

        for item in self.current_result_lines:
            if item.item_type not in {"folder", "file"}:
                continue
            value = item.text.strip()
            if value and MailPrepLogic.is_suspicious_name(value) and value not in seen:
                seen.add(value)
                names.append(value)

        for name in self.current_shot_names:
            if MailPrepLogic.is_suspicious_name(name) and name not in seen:
                seen.add(name)
                names.append(name)

        return names

    def _warning_details(self) -> List[str]:
        warnings = []

        if (
            self.include_subject_check.isChecked()
            and not self.subject_edit.text().strip()
        ):
            warnings.append("Package name is empty")

        if (
            self.include_shot_count_check.isChecked()
            and not self._resolved_shot_count()
        ):
            warnings.append("Shot count empty")

        if (
            self.include_submission_notes_check.isChecked()
            and self.current_shot_names
            and not self._submission_notes_items()
        ):
            warnings.append("Submission notes enabled but empty")

        if self.current_duplicate_shot_names:
            warnings.append(
                f"Duplicate shot names detected ({len(self.current_duplicate_shot_names)})"
            )

        if self.current_missing_summary_count:
            warnings.append(
                f"Missing frames found in {self.current_missing_summary_count} sequence group(s)"
            )

        if self.current_suspicious_names:
            warnings.append(
                f"{len(self.current_suspicious_names)} suspicious name(s) found"
            )

        return warnings

    def _apply_validation_style(self):
        theme = THEMES[self.current_theme]
        warnings = self.current_warning_messages

        if not warnings:
            bg = theme["valid_bg"]
            fg = theme["valid_fg"]
            border = theme["valid_border"]
        else:
            has_errorish = bool(
                self.current_duplicate_shot_names or self.current_suspicious_names
            )
            if has_errorish:
                bg = theme["error_bg"]
                fg = theme["error_fg"]
                border = theme["error_border"]
            else:
                bg = theme["warn_bg"]
                fg = theme["warn_fg"]
                border = theme["warn_border"]

        self.validation_label.setStyleSheet(
            f"background:{bg}; color:{fg}; border:1px solid {border};"
        )

    def _update_validation_summary(self):
        self.current_warning_messages = self._warning_details()

        if not self.current_warning_messages:
            self.validation_label.setText("\u2705 Ready")
            self.validation_label.setToolTip("No warnings found.")
        else:
            warning_count = len(self.current_warning_messages)
            summary = f"\u26a0 {warning_count} warning{'s' if warning_count != 1 else ''}"

            if self.current_suspicious_names:
                summary += f" | {len(self.current_suspicious_names)} suspicious names"
            elif self.current_duplicate_shot_names:
                summary += f" | {len(self.current_duplicate_shot_names)} duplicates"

            self.validation_label.setText(summary)
            self.validation_label.setToolTip("\n".join(self.current_warning_messages))

        self._apply_validation_style()

    # ----------------------------------------------------------
    # Recent paths
    # ----------------------------------------------------------
    def _add_recent_path(self, path: str):
        path = (path or "").strip()
        if not path:
            return
        if path in self.recent_paths:
            self.recent_paths.remove(path)
        self.recent_paths.insert(0, path)
        self.recent_paths = self.recent_paths[:MAX_RECENT_PATHS]
        self._save_recent_paths()
        self._refresh_recent_paths_combo()

    def _load_recent_paths(self):
        try:
            if os.path.isfile(RECENT_PATHS_FILE):
                with open(RECENT_PATHS_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, list):
                    self.recent_paths = [str(x) for x in data if str(x).strip()]
                else:
                    self.recent_paths = []
            else:
                self.recent_paths = []
        except Exception:
            self.recent_paths = []

    def _save_recent_paths(self):
        try:
            with open(RECENT_PATHS_FILE, "w", encoding="utf-8") as f:
                json.dump(self.recent_paths[:MAX_RECENT_PATHS], f, indent=2)
        except Exception:
            pass

    def _refresh_recent_paths_combo(self):
        self.updating_recent_combo = True
        self.recent_paths_combo.clear()
        self.recent_paths_combo.addItem("Recent folders...", "")
        for path in self.recent_paths:
            label = path if len(path) <= 75 else "..." + path[-72:]
            self.recent_paths_combo.addItem(label, path)
        self.recent_paths_combo.setCurrentIndex(0)
        self.updating_recent_combo = False

    def on_recent_path_selected(self, index: int):
        if self.updating_recent_combo:
            return
        if index <= 0:
            return

        path = self.recent_paths_combo.itemData(index)
        if not path:
            return

        self.current_input_mode = "path"
        self.current_manifest_data = None
        self.auto_refresh_checkbox.setEnabled(True)
        self.source_label.setText("Source: Folder Path")
        self.path_edit.setText(path)

        if not os.path.isdir(path):
            QMessageBox.warning(
                self,
                "Path Missing",
                "Selected recent path no longer exists. Path was filled, but preview was not generated.",
            )
            self.set_status("Recent path no longer exists.")
            return

        self.generate_preview(auto=False)

    # ----------------------------------------------------------
    # Session management
    # ----------------------------------------------------------
    def _session_payload(self) -> dict:
        return {
            "schema_version": "1.0",
            "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "path": self.path_edit.text().strip(),
            "subject": self.subject_edit.text(),
            "package_name": self.subject_edit.text(),
            "shot_count": self.shot_count_edit.text(),
            "theme": self.current_theme,
            "current_input_mode": self.current_input_mode,
            "current_manifest_data": (
                self.current_manifest_data
                if self.current_input_mode == "manifest"
                else None
            ),
            "template_sheet_path": self.template_sheet_path,
            "advanced_open": self.advanced_toggle_btn.isChecked(),
            "auto_refresh": self.auto_refresh_checkbox.isChecked(),
            "include_subject": self.include_subject_check.isChecked(),
            "include_shot_count": self.include_shot_count_check.isChecked(),
            "include_submission_notes": self.include_submission_notes_check.isChecked(),
            "include_meta_in_copy": self.include_meta_in_copy_check.isChecked(),
            "complex_mode": self.complex_mode_check.isChecked(),
            "notes": self.get_submission_notes_map(),
            "common_note": self.common_note_edit.text(),
            "notes_search": self.notes_search_edit.text(),
            "recent_paths": self.recent_paths[:MAX_RECENT_PATHS],
        }

    def _apply_session_payload(self, data: dict):
        if not isinstance(data, dict):
            raise ValueError("Invalid session file format.")

        notes_data = data.get("notes", {})
        if not isinstance(notes_data, dict):
            notes_data = {}

        package_name = str(data.get("package_name", data.get("subject", "")))
        self.subject_edit.setText(package_name)
        self.shot_count_edit.setText(str(data.get("shot_count", "")))
        self.common_note_edit.setText(str(data.get("common_note", "")))
        self.notes_search_edit.setText(str(data.get("notes_search", "")))
        self.shot_count_auto_applied = False
        self.shot_count_user_edited = bool(data.get("shot_count", ""))
        self._set_template_sheet_path(str(data.get("template_sheet_path", "")))

        theme = str(data.get("theme", self.current_theme))
        if theme in THEMES:
            self.current_theme = theme

        self.auto_refresh_checkbox.setChecked(bool(data.get("auto_refresh", True)))
        self.include_subject_check.setChecked(bool(data.get("include_subject", True)))
        self.include_shot_count_check.setChecked(
            bool(data.get("include_shot_count", True))
        )
        self.include_submission_notes_check.setChecked(
            bool(data.get("include_submission_notes", True))
        )
        self.include_meta_in_copy_check.setChecked(
            bool(data.get("include_meta_in_copy", True))
        )
        self.complex_mode_check.setChecked(bool(data.get("complex_mode", False)))

        advanced_open = bool(data.get("advanced_open", False))
        self.advanced_toggle_btn.setChecked(advanced_open)
        self.advanced_frame.setVisible(advanced_open)
        self.advanced_toggle_btn.setText(
            "\u25bc Advanced Options" if advanced_open else "\u25b6 Advanced Options"
        )

        loaded_recent = data.get("recent_paths")
        if isinstance(loaded_recent, list):
            self.recent_paths = [str(x) for x in loaded_recent if str(x).strip()][
                :MAX_RECENT_PATHS
            ]
            self._save_recent_paths()
            self._refresh_recent_paths_combo()

        self.apply_theme()

        input_mode = str(data.get("current_input_mode", "path")).strip().lower()
        path = str(data.get("path", "")).strip()

        self.current_manifest_data = None
        self.current_input_mode = "path"
        self.auto_refresh_checkbox.setEnabled(True)
        self.source_label.setText("Source: Folder Path")

        if input_mode == "manifest" and isinstance(
            data.get("current_manifest_data"), dict
        ):
            self.current_manifest_data = data.get("current_manifest_data")
            self.current_input_mode = "manifest"
            self.auto_refresh_checkbox.setEnabled(False)
            self.source_label.setText("Source: Manifest")
            self.start_preview_job(self.current_manifest_data, auto=False)
        else:
            self.path_edit.setText(path)
            if path:
                self._add_recent_path(path)
            if path and os.path.isdir(path):
                self.start_preview_job(path, auto=False)
            else:
                if path:
                    QMessageBox.warning(
                        self,
                        "Path Missing",
                        "Saved session path no longer exists. Other session values were loaded.",
                    )
                    self.set_status("Session loaded, but saved path is missing.")
                self.on_metadata_change()

        def apply_notes_after_render():
            for shot_name, note_text in notes_data.items():
                editor = self.shot_note_edits.get(shot_name)
                if editor is not None:
                    editor.setPlainText(str(note_text))
            self.on_metadata_change()

        QTimer.singleShot(200, apply_notes_after_render)

    # ----------------------------------------------------------
    # File I/O
    # ----------------------------------------------------------
    def import_notes_file(self):
        if not self.current_shot_names:
            QMessageBox.warning(self, "Warning", "Generate preview first.")
            return

        file_path, _ = QFileDialog.getOpenFileName(
            self, "Import Notes", "", "JSON Files (*.json)"
        )
        if not file_path:
            return

        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, dict):
                raise ValueError("Invalid notes file format.")
            updated = 0
            for shot_name, note_text in data.items():
                editor = self.shot_note_edits.get(str(shot_name))
                if editor is not None:
                    editor.setPlainText(str(note_text))
                    updated += 1
            self.on_metadata_change()
            self.set_status(f"Imported notes for {updated} shot(s).")
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Failed to import notes:\n{exc}")
            self.set_status("Notes import failed.")

    def export_notes_file(self):
        if not self.current_shot_names:
            QMessageBox.warning(self, "Warning", "Generate preview first.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Notes", "", "JSON Files (*.json)"
        )
        if not file_path:
            return

        try:
            data = self.get_submission_notes_map()
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=4, ensure_ascii=False)
            self.set_status("Notes exported.")
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Failed to export notes:\n{exc}")
            self.set_status("Notes export failed.")

    def _find_version_files(self, version_name: str) -> List[str]:
        results: List[str] = []
        stack: List[str] = []
        normalized_target = version_name.lower()

        for item in self.current_result_lines:
            if item.item_type == "folder":
                stack = stack[: item.level]
                stack.append(item.text.strip())
                continue
            if item.item_type != "file":
                continue
            stack = stack[: item.level]
            file_text = item.text.strip()
            if file_text.endswith(" (contains)"):
                file_text = file_text[: -len(" (contains)")].rstrip()
            if not file_text:
                continue
            path_parts = [part.lower() for part in stack if part]
            if normalized_target in file_text.lower() or any(
                normalized_target == part or normalized_target in part
                for part in path_parts
            ):
                results.append(file_text)

        return results

    def _frame_range_for_version(self, version_name: str) -> str:
        files = self._find_version_files(version_name)
        if not files:
            return ""
        grouped = MailPrepLogic.group_sequences(files)
        ranges: List[str] = []
        for item in grouped:
            match = re.search(r"(\d+(?:-\d+)?)\.[^.]+$", item)
            if match:
                ranges.append(match.group(1))
            else:
                ranges.append(item)
        return "; ".join(ranges)

    def export_submission_csv(self):
        if not self.current_result_lines:
            QMessageBox.warning(self, "Warning", "Generate preview first.")
            return
        if not self.current_shot_names:
            QMessageBox.warning(
                self, "Warning", "No version names detected. Generate preview first."
            )
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Submission CSV", "", "CSV Files (*.csv)"
        )
        if not file_path:
            return

        try:
            rows = self._build_submission_export_rows()
            with open(file_path, "w", encoding="utf-8", newline="") as f:
                if self.template_sheet_path and os.path.isfile(self.template_sheet_path):
                    headers, seed_rows = self._load_template_headers_and_rows(
                        self.template_sheet_path
                    )
                    if not headers:
                        raise ValueError(
                            "Attached template has no header row. Please select a valid sheet template."
                        )
                    writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
                    writer.writeheader()
                    for idx, row_data in enumerate(rows):
                        out_row = {header: "" for header in headers}
                        if idx < len(seed_rows):
                            seed = seed_rows[idx]
                            for header in headers:
                                if header in seed and seed[header] is not None:
                                    out_row[header] = str(seed[header])
                        for header in headers:
                            normalized = self._normalize_template_header(header)
                            if normalized in row_data:
                                out_row[header] = row_data[normalized]
                        writer.writerow(out_row)
                    self.set_status("Submission CSV exported using attached template.")
                else:
                    writer = csv.writer(f)
                    writer.writerow(
                        ["Version Name", "Start Range", "End Range", "Submission Notes", "Package Name"]
                    )
                    for row_data in rows:
                        writer.writerow(
                            [
                                row_data.get("versionname", ""),
                                row_data.get("startrange", ""),
                                row_data.get("endrange", ""),
                                row_data.get("submissionnotes", ""),
                                row_data.get("packagename", ""),
                            ]
                        )
                    self.set_status("Submission CSV exported.")
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Failed to export submission CSV:\n{exc}")
            self.set_status("Submission CSV export failed.")

    def save_session_file(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Session", "", "JSON Files (*.json)"
        )
        if not file_path:
            return
        try:
            data = self._session_payload()
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=4, ensure_ascii=False)
            self.set_status("Session saved.")
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Failed to save session:\n{exc}")
            self.set_status("Session save failed.")

    def load_session_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Load Session", "", "JSON Files (*.json)"
        )
        if not file_path:
            return
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            self._apply_session_payload(data)
            self.set_status("Session loaded.")
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Failed to load session:\n{exc}")
            self.set_status("Session load failed.")

    def export_mail_html(self):
        if not self.generated_copy_html:
            QMessageBox.warning(self, "Warning", "Generate preview first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Mail as HTML", "", "HTML Files (*.html)"
        )
        if not file_path:
            return
        try:
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(self.generated_copy_html)
            self.set_status("Mail HTML exported.")
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Failed to export HTML:\n{exc}")
            self.set_status("Mail HTML export failed.")

    def export_mail_txt(self):
        if not self.generated_copy_plain_text:
            QMessageBox.warning(self, "Warning", "Generate preview first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Mail as TXT", "", "Text Files (*.txt)"
        )
        if not file_path:
            return
        try:
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(self.generated_copy_plain_text)
            self.set_status("Mail TXT exported.")
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Failed to export TXT:\n{exc}")
            self.set_status("Mail TXT export failed.")

    # ----------------------------------------------------------
    # Preview / clear
    # ----------------------------------------------------------
    def clear_preview(self):
        self.generated_html = ""
        self.generated_plain_text = ""
        self.generated_clipboard_html = ""
        self.generated_preview_html = ""
        self.generated_qc_preview_html = ""
        self.generated_copy_html = ""
        self.generated_copy_plain_text = ""
        self.generated_full_mail_html = ""
        self.generated_full_mail_plain_text = ""
        self.generated_notes_only_html = ""
        self.generated_notes_only_plain_text = ""
        self.last_preview_signature = None
        self.pending_preview_job = None
        self.current_result_lines = []
        self.current_raw_shot_names = []
        self.current_duplicate_shot_names = []
        self.current_suspicious_names = []
        self.current_missing_summary_count = 0
        self.auto_shot_count_value = None
        self.shot_count_auto_applied = False
        self.shot_count_user_edited = False
        self.preview_box.clear()
        self._sync_qc_preview_window()
        self.preview_count_label.setText("Folders: 0   |   Files: 0")
        self.last_refresh_label.setText("Last Refresh: --")
        self.shot_count_edit.setPlaceholderText("Optional")
        self._rebuild_shot_notes_panel([])
        self.source_label.setText(
            "Source: Folder Path"
            if self.current_input_mode == "path"
            else "Source: Manifest"
        )
        self._update_validation_summary()

    def clear_form(self):
        self.path_edit.clear()
        self.subject_edit.clear()
        self.shot_count_edit.clear()
        self.common_note_edit.clear()
        self.notes_search_edit.clear()
        self.shot_count_auto_applied = False
        self.shot_count_user_edited = False
        self.current_input_mode = "path"
        self.current_manifest_data = None
        self.auto_refresh_checkbox.setEnabled(True)
        self.source_label.setText("Source: Folder Path")
        self.clear_preview()
        self.set_status("Form cleared.")

    def render_preview(self, result: BuildResult):
        existing_notes = self.get_submission_notes_map()

        self.generated_plain_text = result.plain_text
        self.generated_html = result.html_text
        self.generated_clipboard_html = result.clipboard_html
        self.last_preview_signature = result.signature
        self.current_result_lines = result.lines[:]

        shot_names = self._extract_shot_names_from_lines(result.lines)
        self.current_duplicate_shot_names = self._duplicate_shot_names_from_raw(
            self.current_raw_shot_names
        )
        self.current_missing_summary_count = sum(
            1 for item in result.lines if item.item_type == "missing"
        )
        self.auto_shot_count_value = len(shot_names) if shot_names else None
        self.current_suspicious_names = self._collect_suspicious_names()

        # Only auto-fill shot count if the user has not manually edited it
        if self.auto_shot_count_value:
            if not self.shot_count_user_edited:
                self.shot_count_edit.blockSignals(True)
                self.shot_count_edit.setText(str(self.auto_shot_count_value))
                self.shot_count_edit.blockSignals(False)
                self.shot_count_auto_applied = True
        else:
            if self.shot_count_auto_applied and not self.shot_count_user_edited:
                self.shot_count_edit.blockSignals(True)
                self.shot_count_edit.clear()
                self.shot_count_edit.blockSignals(False)
                self.shot_count_auto_applied = False

        self._rebuild_shot_notes_panel(shot_names)

        for shot_name, text in existing_notes.items():
            editor = self.shot_note_edits.get(shot_name)
            if editor is not None and text:
                editor.setPlainText(text)

        self.generated_preview_html = self._compose_preview_html(result.html_text)
        self.generated_qc_preview_html = self._compose_qc_preview_html(result.html_text)
        self.generated_copy_html = self._compose_copy_rich_html(result.clipboard_html)
        self.generated_copy_plain_text = self._compose_copy_plain_text(result.plain_text)
        self.generated_notes_only_html = self._compose_notes_only_html()
        self.generated_notes_only_plain_text = self._compose_notes_only_plain_text()
        self.generated_full_mail_html = self._compose_full_mail_html()
        self.generated_full_mail_plain_text = self._compose_full_mail_plain_text()

        self.preview_box.setHtml(self.generated_preview_html)
        self._sync_qc_preview_window()
        self.preview_count_label.setText(
            f"Folders: {result.folder_count}   |   Files: {result.file_count}"
        )
        self._update_last_refresh()

        self.source_label.setText(
            "Source: Folder Path" if self.current_input_mode == "path" else "Source: Manifest"
        )

        if self.current_input_mode == "path":
            path = self.path_edit.text().strip()
            if path:
                self._add_recent_path(path)

        self._update_validation_summary()

    # ----------------------------------------------------------
    # Event handlers
    # ----------------------------------------------------------
    def on_path_change(self):
        if self.current_input_mode != "path":
            self.current_input_mode = "path"
            self.current_manifest_data = None
            self.auto_refresh_checkbox.setEnabled(True)
            self.source_label.setText("Source: Folder Path")
        self.path_debounce_timer.start(100)

    def on_shot_count_edited(self, _text: str):
        """Called only when the user manually types in the shot count field."""
        self.shot_count_auto_applied = False
        self.shot_count_user_edited = True

    def on_metadata_change(self):
        """Debounced — triggers _do_metadata_update after a short delay."""
        self.metadata_debounce_timer.start(120)

    def _do_metadata_update(self):
        """Actual metadata re-render, called after debounce settles."""
        if not self.generated_plain_text:
            self.filter_notes_panel()
            self._update_validation_summary()
            return

        self.current_duplicate_shot_names = self._duplicate_shot_names_from_raw(
            self.current_raw_shot_names
        )
        self.current_suspicious_names = self._collect_suspicious_names()
        self._refresh_note_label_styles()

        self.generated_preview_html = self._compose_preview_html(self.generated_html)
        self.generated_qc_preview_html = self._compose_qc_preview_html(self.generated_html)
        self.generated_copy_html = self._compose_copy_rich_html(self.generated_clipboard_html)
        self.generated_copy_plain_text = self._compose_copy_plain_text(self.generated_plain_text)
        self.generated_notes_only_html = self._compose_notes_only_html()
        self.generated_notes_only_plain_text = self._compose_notes_only_plain_text()
        self.generated_full_mail_html = self._compose_full_mail_html()
        self.generated_full_mail_plain_text = self._compose_full_mail_plain_text()
        self.preview_box.setHtml(self.generated_preview_html)
        self._sync_qc_preview_window()

        existing_notes = self.get_submission_notes_map()
        shot_names = self._extract_shot_names_from_lines(self.current_result_lines)
        if shot_names != self.current_shot_names:
            self.current_duplicate_shot_names = self._duplicate_shot_names_from_raw(
                self.current_raw_shot_names
            )
            self._rebuild_shot_notes_panel(shot_names)
            for shot_name, text in existing_notes.items():
                editor = self.shot_note_edits.get(shot_name)
                if editor is not None and text:
                    editor.setPlainText(text)

            if self.auto_shot_count_value and not self.shot_count_user_edited:
                self.shot_count_edit.blockSignals(True)
                self.shot_count_edit.setText(str(self.auto_shot_count_value))
                self.shot_count_edit.blockSignals(False)
                self.shot_count_auto_applied = True

            self.current_suspicious_names = self._collect_suspicious_names()

            # Recompose after shot panel rebuild
            self.generated_preview_html = self._compose_preview_html(self.generated_html)
            self.generated_qc_preview_html = self._compose_qc_preview_html(self.generated_html)
            self.generated_copy_html = self._compose_copy_rich_html(self.generated_clipboard_html)
            self.generated_copy_plain_text = self._compose_copy_plain_text(self.generated_plain_text)
            self.generated_notes_only_html = self._compose_notes_only_html()
            self.generated_notes_only_plain_text = self._compose_notes_only_plain_text()
            self.generated_full_mail_html = self._compose_full_mail_html()
            self.generated_full_mail_plain_text = self._compose_full_mail_plain_text()
            self.preview_box.setHtml(self.generated_preview_html)
            self._sync_qc_preview_window()

        self.filter_notes_panel()
        self._update_validation_summary()

    def on_auto_refresh_toggle(self, checked: bool):
        self.hint_label.setText(
            "Auto-refresh is ON" if checked else "Auto-refresh is OFF"
        )

    def toggle_advanced_options(self):
        is_open = self.advanced_toggle_btn.isChecked()
        self.advanced_frame.setVisible(is_open)
        self.advanced_toggle_btn.setText(
            "\u25bc Advanced Options" if is_open else "\u25b6 Advanced Options"
        )

    def toggle_qc_preview_mode(self):
        self.qc_preview_mode = self.preview_dialog_btn.isChecked()
        self.preview_dialog_btn.setText(
            "\U0001f5bc QC Preview: ON" if self.qc_preview_mode else "\U0001f5bc QC Preview: OFF"
        )

        if self.qc_preview_mode:
            if not self.generated_qc_preview_html:
                QMessageBox.warning(self, "Warning", "Generate preview first.")
                self.preview_dialog_btn.blockSignals(True)
                self.preview_dialog_btn.setChecked(False)
                self.preview_dialog_btn.blockSignals(False)
                self.qc_preview_mode = False
                self.preview_dialog_btn.setText("\U0001f5bc QC Preview: OFF")
                return
            self._show_qc_preview_window()
            self.set_status("QC preview mode enabled.")
        else:
            self._close_qc_preview_window()

    def _debounced_path_preview(self):
        if self.current_input_mode != "path":
            return
        path = self.path_edit.text().strip()
        if not path:
            self.clear_preview()
            self.set_status("Ready.")
            return
        if os.path.isdir(path):
            self.start_preview_job(path, auto=True)
        else:
            self.clear_preview()
            self.set_status("Waiting for valid folder path...")

    def auto_refresh_tick(self):
        if self.current_input_mode != "path":
            return
        if not self.auto_refresh_checkbox.isChecked():
            return
        path = self.path_edit.text().strip()
        if (
            path
            and os.path.isdir(path)
            and self.generated_plain_text
            and not self.loading_active
        ):
            current_sig = MailPrepLogic.path_signature(path)
            if current_sig != self.last_preview_signature:
                self.start_preview_job(path, auto=True)

    # ----------------------------------------------------------
    # Job management
    # ----------------------------------------------------------
    def start_preview_job(self, source_data, auto: bool = False):
        if self.loading_active:
            self.pending_preview_job = (self.current_input_mode, source_data, auto)
            return

        self.pending_preview_job = None

        if self.current_input_mode == "path":
            path = str(source_data).strip() if source_data is not None else ""
            if not path:
                if not auto:
                    QMessageBox.warning(self, "Warning", "Please select a folder first.")
                self.clear_preview()
                self.set_status("Select a folder first.")
                return
            if not os.path.isdir(path):
                if not auto:
                    QMessageBox.critical(self, "Error", "Selected path is not a valid folder.")
                self.clear_preview()
                self.set_status("Invalid folder path.")
                return

        elif self.current_input_mode == "manifest":
            if not source_data:
                QMessageBox.warning(self, "Warning", "Load a manifest first.")
                self.clear_preview()
                self.set_status("Load a manifest first.")
                return

        self.current_job_id += 1
        job_id = self.current_job_id

        self.set_loading(True, "Scanning folder...")
        self.set_status(
            "Generating preview..." if not auto else "Auto refreshing preview..."
        )

        self.worker_thread = QThread()
        self.worker = PreviewWorker(self.current_input_mode, source_data, self.current_theme)
        self.worker.moveToThread(self.worker_thread)

        self.worker_thread.started.connect(self.worker.run)
        self.worker.finished.connect(
            lambda result: self.finish_preview_job(result, source_data, job_id, auto)
        )
        self.worker.failed.connect(
            lambda error: self.fail_preview_job(error, source_data, job_id, auto)
        )

        self.worker.finished.connect(self.worker_thread.quit)
        self.worker.failed.connect(self.worker_thread.quit)
        self.worker.finished.connect(self.worker.deleteLater)
        self.worker.failed.connect(self.worker.deleteLater)
        self.worker_thread.finished.connect(self.worker_thread.deleteLater)

        self.worker_thread.start()

    def _run_pending_preview_job_if_any(self):
        if self.loading_active or not self.pending_preview_job:
            return
        pending_mode, pending_source_data, pending_auto = self.pending_preview_job
        self.pending_preview_job = None

        if pending_mode != self.current_input_mode:
            if self.current_input_mode == "path":
                pending_source_data = self.path_edit.text().strip()
            elif self.current_input_mode == "manifest":
                pending_source_data = self.current_manifest_data
            else:
                pending_source_data = None

        self.start_preview_job(pending_source_data, auto=pending_auto)

    def finish_preview_job(
        self, result: BuildResult, source_data, job_id: int, auto: bool
    ):
        if job_id != self.current_job_id:
            return

        self.set_loading(False)

        if self.current_input_mode == "path":
            current_path = self.path_edit.text().strip()
            source_path = str(source_data).strip() if source_data is not None else ""
            if current_path != source_path:
                self._run_pending_preview_job_if_any()
                return

        try:
            self.render_preview(result)
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Failed to render preview:\n{exc}")
            self.set_status("Preview render failed.")
            self._run_pending_preview_job_if_any()
            return

        self.set_status(
            "Preview auto-refreshed." if auto else "Preview generated successfully."
        )
        self._run_pending_preview_job_if_any()

    def fail_preview_job(self, error: str, source_data, job_id: int, auto: bool):
        if job_id != self.current_job_id:
            return
        self.set_loading(False)
        if not auto:
            QMessageBox.critical(self, "Error", f"Failed to generate preview:\n{error}")
        self.set_status(f"Error while generating preview: {error}")
        self._run_pending_preview_job_if_any()

    def generate_preview(self, auto: bool = False):
        if self.current_input_mode == "path":
            source = self.path_edit.text().strip()
        elif self.current_input_mode == "manifest":
            source = self.current_manifest_data
        else:
            source = None
        self.start_preview_job(source, auto=auto)

    # ----------------------------------------------------------
    # Copy actions
    # ----------------------------------------------------------
    def copy_rich_text(self):
        if not self.generated_copy_html:
            QMessageBox.warning(self, "Warning", "Generate preview first.")
            return
        try:
            self._copy_html_to_clipboard(self.generated_copy_html, self.generated_copy_plain_text)
            self._set_copy_success(self.copy_rich_btn, "\u2705 Rich Text Copied", "Rich text copied.")
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Copy failed:\n{exc}")
            self.set_status("Rich text copy failed.")

    def copy_html_source(self):
        if not self.generated_copy_html:
            QMessageBox.warning(self, "Warning", "Generate preview first.")
            return
        try:
            clipboard = QGuiApplication.clipboard()
            clipboard.setText(self.generated_copy_html, QClipboard.Clipboard)
            try:
                clipboard.setText(self.generated_copy_html, QClipboard.Selection)
            except Exception:
                pass
            self._set_copy_success(self.copy_html_btn, "\u2705 HTML Copied", "HTML source copied.")
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Copy failed:\n{exc}")
            self.set_status("HTML source copy failed.")

    def copy_full_mail(self):
        if not self.generated_full_mail_html:
            QMessageBox.warning(self, "Warning", "Generate preview first.")
            return
        try:
            self._copy_html_to_clipboard(self.generated_full_mail_html, self.generated_full_mail_plain_text)
            self._set_copy_success(self.copy_full_mail_btn, "\u2705 Full Mail Copied", "Full mail body copied.")
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Copy failed:\n{exc}")
            self.set_status("Full mail copy failed.")

    def copy_full_mail_html(self):
        if not self.generated_full_mail_html:
            QMessageBox.warning(self, "Warning", "Generate preview first.")
            return
        try:
            clipboard = QGuiApplication.clipboard()
            clipboard.setText(self.generated_full_mail_html, QClipboard.Clipboard)
            try:
                clipboard.setText(self.generated_full_mail_html, QClipboard.Selection)
            except Exception:
                pass
            self._set_copy_success(
                self.copy_full_mail_html_btn, "\u2705 Full Mail HTML Copied", "Full mail HTML copied."
            )
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Copy failed:\n{exc}")
            self.set_status("Full mail HTML copy failed.")

    def copy_notes_only(self):
        if not self.generated_notes_only_html:
            QMessageBox.warning(self, "Warning", "No notes available to copy.")
            return
        try:
            self._copy_html_to_clipboard(
                self.generated_notes_only_html, self.generated_notes_only_plain_text
            )
            self.set_status("Submission notes copied.")
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Copy failed:\n{exc}")
            self.set_status("Notes copy failed.")

    # ----------------------------------------------------------
    # Template sheet
    # ----------------------------------------------------------
    def _set_template_sheet_path(self, path: str):
        path = (path or "").strip()
        if path and not os.path.isfile(path):
            path = ""
        self.template_sheet_path = path
        if not path:
            self.attach_sheet_btn.setText("\U0001f4ce Attach Sheet")
            self.attach_sheet_btn.setToolTip("Attach a client sheet template")
            return
        base_name = os.path.basename(path)
        button_label = base_name if len(base_name) <= 28 else f"{base_name[:25]}..."
        self.attach_sheet_btn.setText(f"\U0001f4ce {button_label}")
        self.attach_sheet_btn.setToolTip(path)

    def attach_template_sheet(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Attach Sheet Template",
            "",
            "Template Files (*.csv *.xlsx *.xls);;CSV Files (*.csv);;Excel Files (*.xlsx *.xls)",
        )
        if not file_path:
            return
        ext = os.path.splitext(file_path)[1].lower()
        if ext not in TEMPLATE_SHEET_EXTENSIONS:
            QMessageBox.warning(self, "Unsupported Template", "Please select a CSV, XLSX, or XLS template file.")
            return
        self._set_template_sheet_path(file_path)
        self.set_status(f"Template attached: {os.path.basename(file_path)}")

    def _normalize_template_header(self, header: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", (header or "").strip().lower())

    def _load_template_headers_and_rows(
        self, template_path: str
    ) -> Tuple[List[str], List[Dict[str, str]]]:
        ext = os.path.splitext(template_path)[1].lower()

        if ext == ".csv":
            with open(template_path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                headers = [h for h in (reader.fieldnames or []) if h is not None]
                rows = []
                for row in reader:
                    rows.append(
                        {
                            key: ("" if value is None else str(value))
                            for key, value in row.items()
                            if key is not None
                        }
                    )
                return headers, rows

        if ext == ".xlsx":
            try:
                from openpyxl import load_workbook
            except Exception as exc:
                raise RuntimeError("XLSX template requires openpyxl. Install it and retry.") from exc

            workbook = load_workbook(template_path, data_only=True, read_only=True)
            try:
                sheet = workbook.active
                values = list(sheet.iter_rows(values_only=True))
            finally:
                workbook.close()

            if not values:
                return [], []

            headers = ["" if v is None else str(v).strip() for v in values[0]]
            rows = []
            for raw in values[1:]:
                row = {}
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    value = raw[idx] if idx < len(raw) else ""
                    row[header] = "" if value is None else str(value)
                if any(str(v).strip() for v in row.values()):
                    rows.append(row)
            return [h for h in headers if h], rows

        if ext == ".xls":
            try:
                import xlrd
            except Exception as exc:
                raise RuntimeError(
                    "XLS template requires xlrd. Install it and retry, or convert template to XLSX/CSV."
                ) from exc

            workbook = xlrd.open_workbook(template_path)
            sheet = workbook.sheet_by_index(0)
            if sheet.nrows <= 0:
                return [], []

            headers = [str(sheet.cell_value(0, col)).strip() for col in range(sheet.ncols)]
            rows = []
            for row_idx in range(1, sheet.nrows):
                row = {}
                for col_idx, header in enumerate(headers):
                    if not header:
                        continue
                    row[header] = str(sheet.cell_value(row_idx, col_idx)).strip()
                if any(str(v).strip() for v in row.values()):
                    rows.append(row)
            return [h for h in headers if h], rows

        raise ValueError("Unsupported template format.")

    def _build_submission_export_rows(self) -> List[Dict[str, str]]:
        rows: List[Dict[str, str]] = []
        notes_map = self.get_submission_notes_map()
        package_name = self.subject_edit.text().strip()
        shot_count = self._resolved_shot_count()

        for version_name in self.current_shot_names:
            frame_range_str = self._frame_range_for_version(version_name)
            start_range = ""
            end_range = ""

            ranges = frame_range_str.split("; ")
            if ranges and ranges[0].strip():
                first_range = ranges[0].strip()
                if "-" in first_range:
                    parts = first_range.split("-", 1)
                    start_range = parts[0].strip()
                    end_range = parts[1].strip()
                else:
                    start_range = first_range.strip()
                    end_range = start_range

            note_value = notes_map.get(version_name, "")
            normalized_values = {
                "versionname": version_name,
                "version": version_name,
                "shotname": version_name,
                "shot": version_name,
                "startrange": start_range,
                "startframe": start_range,
                "in": start_range,
                "endrange": end_range,
                "endframe": end_range,
                "out": end_range,
                "framerange": frame_range_str,
                "submissionnotes": note_value,
                "notes": note_value,
                "note": note_value,
                "comments": note_value,
                "packagename": package_name,
                "package": package_name,
                "subject": package_name,
                "shotcount": shot_count,
            }
            rows.append(normalized_values)

        return rows

    # ----------------------------------------------------------
    # Folder / manifest
    # ----------------------------------------------------------
    def select_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Folder")
        if folder:
            self.current_input_mode = "path"
            self.current_manifest_data = None
            self.auto_refresh_checkbox.setEnabled(True)
            self.source_label.setText("Source: Folder Path")
            self.path_edit.setText(folder)
            self._add_recent_path(folder)

    def load_manifest_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Load Manifest", "", "JSON Files (*.json)"
        )
        if not file_path:
            return
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, dict) or "items" not in data:
                raise ValueError("Invalid manifest format. Missing 'items'.")
            self.current_manifest_data = data
            self.current_input_mode = "manifest"
            self.auto_refresh_checkbox.setEnabled(False)
            self.source_label.setText("Source: Manifest")
            self.start_preview_job(data, auto=False)
            self.set_status("Manifest loaded successfully.")
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Failed to load manifest:\n{exc}")
            self.set_status("Manifest load failed.")

    def export_manifest_file(self):
        if not self.current_result_lines:
            QMessageBox.warning(self, "Warning", "Generate preview first.")
            return
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Manifest", "", "JSON Files (*.json)"
        )
        if not file_path:
            return
        try:
            data = MailPrepLogic.lines_to_manifest(self.current_result_lines)
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=4)
            self.set_status("Manifest exported.")
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Failed to export manifest:\n{exc}")
            self.set_status("Manifest export failed.")

    # ----------------------------------------------------------
    # Warnings / preview dialog
    # ----------------------------------------------------------
    def show_warnings(self):
        try:
            warnings = self._warning_details()
            if warnings:
                message = []
                for idx, warning in enumerate(warnings, start=1):
                    message.append(f"{idx}. {warning}")
                if self.current_suspicious_names:
                    message.append("\nSuspicious names:")
                    for name in self.current_suspicious_names:
                        message.append(f"   - {name}")
                QMessageBox.information(self, "Validation Warnings", "\n".join(message))
            else:
                QMessageBox.information(self, "Validation", "No issues found.")
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Failed to show warnings:\n{exc}")

    def open_preview_dialog(self):
        html_to_show = self.generated_preview_html
        if not html_to_show:
            QMessageBox.warning(self, "Warning", "Generate preview first.")
            return
        dialog = PreviewDialog(self, html_content=html_to_show, theme_name=self.current_theme)
        dialog.exec_()

    # ----------------------------------------------------------
    # Close
    # ----------------------------------------------------------
    def closeEvent(self, event):
        try:
            self.path_debounce_timer.stop()
            self.metadata_debounce_timer.stop()
            self.auto_refresh_timer.stop()
            self.copy_feedback_timer.stop()
            if self.qc_preview_window is not None:
                self.qc_preview_window.close()
            if self.worker is not None:
                self.worker.cancel()
            if self.worker_thread is not None and self.worker_thread.isRunning():
                self.worker_thread.quit()
                if not self.worker_thread.wait(2000):
                    self.worker_thread.terminate()
                    self.worker_thread.wait(1000)
        except Exception:
            pass
        super().closeEvent(event)


# ------------------------------------------------------------
# Regression tests
# ------------------------------------------------------------
def _run_shot_extraction_regression_tests():
    app = QApplication.instance() or QApplication(sys.argv)
    window = MailPrepWindow()
    window.complex_mode_check.setChecked(True)

    lines = [
        LineItem("folder", 0, "prl5780"),
        LineItem("folder", 1, "Rotomation"),
        LineItem("file", 2, "manifest.yaml"),
        LineItem("folder", 2, "maya"),
        LineItem("file", 3, "prl5780_mp1_v1_retimed_main_rotomation_botvfx_v1.mb"),
        LineItem("folder", 2, "nuke"),
        LineItem("file", 3, "prl5780_mp1_v1_retimed_main_rotomation_botvfx_v1.nk"),
        LineItem("folder", 0, "spi2875"),
        LineItem("folder", 1, "splines"),
        LineItem("file", 2, "spi2875.mp01.v0001.SRC.roto.v02.sfx"),
    ]

    extracted = window._extract_shot_names_from_lines(lines)
    expected = [
        "prl5780_mp1_v1_retimed_main_rotomation_botvfx_v1",
        "spi2875.mp01.v0001.SRC.roto.v02",
    ]
    if extracted != expected:
        raise AssertiоnError(
            f"Shot extraction regression failed: expected {expected}, got {extracted}"
        )
    print("Shot extraction regression test passed")


def main():
    if os.environ.get("MAILPREP_RUN_REGRESSION_TESTS") == "1":
        _run_shot_extraction_regression_tests()

    app = QApplication(sys.argv)
    window = MailPrepWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
